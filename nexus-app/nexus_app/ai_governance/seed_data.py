"""Parse governance rules and tag dimensions from Excel and produce seed data.

Reads ``docs/ai-governance/20260605数据清单.xlsx`` and builds the complete
``rules_content`` dict for the first ``GovernanceRulesVersion`` (version=1,
status=active).  This is the **new** source of truth — ``config/governance_rules.json``
is deprecated.

Sheets parsed
-------------
* **对应分类说明** — 18 columns (A-R), 2 header rows, N data rows.
  Produces the ``classifications`` list with 5-dimension tagging basis,
  quality scoring, and chunking metadata per classification.
* **标签维度选项** — 5 columns (A-E), 2 header rows.
  Produces the ``tag_dimensions`` top-level key.

Column mapping (对应分类说明)
-----------------------------
===== ===================== ==================================================
Col   Field                 Stage / Purpose
===== ===================== ==================================================
A     parent_type           父类型（行业、产业数据 / 岗位&职业数据 / 专业数据）
B     name                  子类型名称
C     description           文档分类说明
D     application_scenarios 应用场景（后期上游消费，14 类，不参与治理决策）
E     title_keywords        AI 分类判定——标题关键词（; 分隔）
F     content_keywords      AI 分类判定——内容关键词（; 分隔）
G     tagging_basis         标签打标——专业领域维度有效值+打标依据
H     education_level       标签打标——学历层次（分类特定）
I     geo_scope             标签打标——地域范围规则
J     timeliness            标签打标——时效性维度有效值+打标依据
K     data_source           标签打标——数据来源维度有效值+打标依据
L     quality_reliability   质量评分——来源可靠性
M     quality_timeliness    质量评分——时效性
N     quality_completeness  质量评分——完整性
O     quality_other         质量评分——其他维度
P     decomposition_note    后期 chunking——是否结构清晰
Q     structure_description 后期 chunking——文档结构说明
R     remarks               备注
===== ===================== ==================================================
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Default repo-relative path
# ---------------------------------------------------------------------------

_DEFAULT_EXCEL_PATH = (
    Path(__file__).resolve().parent.parent.parent.parent
    / "docs" / "ai-governance" / "20260605数据清单.xlsx"
)

# ---------------------------------------------------------------------------
# Built-in levels (L1-L4) — fundamental to the system, not defined in Excel
# ---------------------------------------------------------------------------

_BUILTIN_LEVELS: list[dict[str, Any]] = [
    {
        "code": "L1",
        "name": "公开",
        "description": "可对外公开的数据，无敏感性限制",
        "criteria": [
            "数据内容已经过审批可对外发布",
            "不包含个人隐私、商业机密或敏感业务信息",
            "可在公开渠道分享和传播",
        ],
        "requires_approval": False,
    },
    {
        "code": "L2",
        "name": "内部",
        "description": "仅限内部员工访问的数据，不得对外分享",
        "criteria": [
            "数据仅供企业内部使用",
            "包含内部流程、规范或一般性业务信息",
            "需要内部身份认证才能访问",
            "不包含高度敏感的个人信息或核心商业机密",
        ],
        "requires_approval": False,
    },
    {
        "code": "L3",
        "name": "机密",
        "description": "敏感业务数据，需要明确授权才能访问",
        "criteria": [
            "包含敏感个人信息（如绩效、薪资、医疗等）",
            "包含核心商业机密、竞争策略、未公开财务数据",
            "包含重要合同条款、客户核心信息",
            "需要明确的访问授权和审计记录",
            "外部AI模型不得接收未脱敏的L3内容",
        ],
        "requires_approval": True,
        "forbid_external_llm": True,
    },
    {
        "code": "L4",
        "name": "绝密",
        "description": "最高敏感级别数据，严格控制访问",
        "criteria": [
            "包含国家秘密、监管要求保护的数据",
            "包含最高级别商业机密、核心技术秘密",
            "包含重大未公开事项、重要人事决策",
            "必须经过最高级别审批才能访问",
            "禁止传输到任何外部系统或AI模型",
        ],
        "requires_approval": True,
        "forbid_external_llm": True,
    },
]

# ---------------------------------------------------------------------------
# Default quality scoring configuration
# ---------------------------------------------------------------------------

_DEFAULT_QUALITY_SCORING: dict[str, Any] = {
    "dimensions": [
        {
            "name": "completeness",
            "weight": 0.30,
            "description": "内容完整性：必要字段是否完整",
            "check_items": [
                {"name": "has_title", "description": "文档必须有标题", "severity": "blocking"},
                {"name": "has_content", "description": "文档必须有实质性内容", "severity": "blocking"},
            ],
        },
        {
            "name": "accuracy",
            "weight": 0.25,
            "description": "内容准确性：分类建议与内容的匹配程度",
            "check_items": [
                {"name": "classification_confidence", "description": "分类置信度必须达到阈值", "severity": "warning"},
                {"name": "no_conflicting_signals", "description": "分类和分级建议不应有矛盾信号", "severity": "blocking"},
            ],
        },
        {
            "name": "consistency",
            "weight": 0.25,
            "description": "一致性：与同类资产的分类分级是否一致",
            "check_items": [
                {"name": "level_matches_classification", "description": "分级应与分类对应的默认级别范围一致", "severity": "warning"},
                {"name": "org_scope_valid", "description": "组织范围引用必须是有效的组织单元", "severity": "blocking"},
            ],
        },
        {
            "name": "usability",
            "weight": 0.20,
            "description": "可用性：内容是否适合被检索、问答和业务使用",
            "check_items": [
                {"name": "content_length_adequate", "description": "内容长度应满足最低要求", "severity": "warning"},
                {"name": "no_parse_errors", "description": "解析结果不应有严重错误", "severity": "blocking"},
            ],
        },
    ],
    "thresholds": {"pass": 80, "warning": 60, "review_required_below": 50},
    "confidence_threshold_auto_adopt": 0.85,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _split_keywords(value: str | None) -> list[str]:
    """Split on ``;``, ``；``, or ``、`` (enumeration comma) into deduplicated list."""
    if not value or not value.strip() or value.strip() == "/":
        return []
    parts = re.split(r"[;；、]", value)
    return [p.strip() for p in parts if p.strip()]


def _parse_tag_dimension_values(text: str | None) -> list[dict[str, str]]:
    r"""Parse ``值（依据：...）；值（依据：...）`` into structured list.

    Returns ``[{"value": "...", "criteria": "..."}, ...]``.

    Falls back to a single description entry if the text is prose (no
    structured ``值（依据：...）`` pattern detected).
    """
    if not text or not text.strip() or text.strip() == "/":
        return []
    result: list[dict[str, str]] = []
    # Split on ；or ; then try to parse each entry as "值（依据：...）"
    entries = re.split(r"[;；]\s*", text.strip())
    # If there are very few semicolons and no structured pattern, treat as prose
    structured_count = 0
    for entry in entries:
        entry = entry.strip().rstrip("；;")
        if not entry:
            continue
        if re.match(r"^.+?[（(]依据[:：].+?[）)]$", entry):
            structured_count += 1
    # If no structured entries found, return the whole text as one prose description
    if structured_count == 0:
        return [{"value": text.strip(), "criteria": ""}]
    for entry in entries:
        entry = entry.strip().rstrip("；;")
        if not entry:
            continue
        m = re.match(r"^(.+?)[（(]依据[:：]\s*(.+?)[）)]$", entry)
        if m:
            result.append({"value": m.group(1).strip(), "criteria": m.group(2).strip()})
        else:
            result.append({"value": entry, "criteria": ""})
    return result


def _build_code(name: str) -> str:
    """Derive a machine-friendly code from a Chinese classification name."""
    _MAP: dict[str, str] = {
        "产业政策": "industry_policy",
        "产业报告": "industry_report",
        "行业报告": "sector_report",
        "岗位需求数据": "job_demand",
        "职业能力分析表": "competency_analysis",
        "职业类证书": "vocational_certificate",
        "专业教学标准": "teaching_standard",
        "专业布点数": "program_distribution",
        "专业人才需求报告": "talent_demand_report",
        "人才培养方案": "talent_training_plan",
        "专业简介": "program_profile",
    }
    return _MAP.get(name, re.sub(r"[^\w]", "_", name).strip("_").lower())


def _clean_cell(value: Any) -> str | None:
    """Return a cleaned string from a cell, or None if empty/invalid."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s == "/":
        return None
    return s


# ---------------------------------------------------------------------------
# Sheet parsers
# ---------------------------------------------------------------------------

def parse_tag_dimensions(excel_path: str | Path | None = None) -> dict[str, Any]:
    """Parse Sheet ``标签维度选项`` → ``tag_dimensions`` dict.

    Returns the 5-dimension definition with:
    - 4 ``per_classification`` dimensions (values come from 对应分类说明 columns G/I/J/K)
    - 1 ``fixed`` dimension (学历层次)
    """
    import openpyxl

    path = Path(excel_path) if excel_path else _DEFAULT_EXCEL_PATH
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb["标签维度选项"]

    _rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

    education_values: list[str] = []
    for row in _rows:
        val = _clean_cell(row[1])  # Column B
        if val and val not in education_values:
            education_values.append(val)

    wb.close()

    return {
        "professional_domain": {
            "name": "专业领域",
            "source": "per_classification",
            "description": "从各分类 G 列（标签维度及打标依据）中按分类提取",
            "binding_column": "tagging_basis",
        },
        "education_level": {
            "name": "学历层次",
            "source": "fixed",
            "values": education_values,
        },
        "geographic_scope": {
            "name": "地域范围",
            "source": "per_classification",
            "description": "按发布主体、适用范围、地区名称打标",
            "binding_column": "geo_scope",
        },
        "timeliness": {
            "name": "时效性",
            "source": "per_classification",
            "binding_column": "timeliness",
        },
        "data_source_type": {
            "name": "数据来源",
            "source": "per_classification",
            "binding_column": "data_source",
        },
    }


def parse_classifications(excel_path: str | Path | None = None) -> list[dict[str, Any]]:
    """Parse Sheet ``对应分类说明`` → list of extended classification dicts.

    Each dict contains the full 18-column mapping including title/content
    keywords for AI classification, 5-dimension tagging basis, per-classification
    quality dimensions, and chunking metadata.
    """
    import openpyxl

    path = Path(excel_path) if excel_path else _DEFAULT_EXCEL_PATH
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb["对应分类说明"]

    rows = list(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True))

    classifications: list[dict[str, Any]] = []
    current_parent: str | None = None

    for row in rows:
        # Unpack 18 columns
        parent_cell = _clean_cell(row[0])   # A
        name        = _clean_cell(row[1])   # B
        description = _clean_cell(row[2])   # C
        scenarios   = _clean_cell(row[3])   # D
        title_kw    = _clean_cell(row[4])   # E
        content_kw  = _clean_cell(row[5])   # F
        prof_domain = _clean_cell(row[6])   # G
        edu_level   = _clean_cell(row[7])   # H
        geo_scope   = _clean_cell(row[8])   # I
        timeliness  = _clean_cell(row[9])   # J
        data_src    = _clean_cell(row[10])  # K
        q_relia     = _clean_cell(row[11])  # L
        q_time      = _clean_cell(row[12])  # M
        q_compl     = _clean_cell(row[13])  # N
        q_other     = _clean_cell(row[14])  # O
        decomp_note = _clean_cell(row[15])  # P
        structure   = _clean_cell(row[16])  # Q
        # Col R (remarks) is ignored

        # Update parent type when column A is present
        if parent_cell:
            current_parent = parent_cell

        # Skip truly empty rows
        if name is None:
            continue

        code = _build_code(name)

        # --- Build quality dimensions from L-O columns ---
        quality_dims: list[dict[str, Any]] = []
        if q_relia:
            quality_dims.append(_parse_quality_dim(q_relia, "来源可靠性"))
        if q_time:
            quality_dims.append(_parse_quality_dim(q_time, "信息时效性"))
        if q_compl:
            quality_dims.append(_parse_quality_dim(q_compl, "内容完整性"))
        if q_other:
            quality_dims.append(_parse_quality_dim(q_other, "其他维度"))

        # --- Build classification dict ---
        cls_dict: dict[str, Any] = {
            "code": code,
            "name": name,
            "parent_type": current_parent,
            "description": description or "",
            "application_scenarios": _split_keywords(scenarios),
            # AI classification stage (E/F columns)
            "title_keywords": _split_keywords(title_kw),
            "content_keywords": _split_keywords(content_kw),
            # AI tagging stage (G-K columns)
            "tagging_basis": {
                "professional_domain": _parse_tag_dimension_values(prof_domain),
                "education_level": _clean_cell(edu_level) or "",
            },
            "geo_scope": _clean_cell(geo_scope) or "",
            "timeliness": _parse_tag_dimension_values(timeliness),
            "data_source": _parse_tag_dimension_values(data_src),
            # Quality scoring (L-O columns)
            "quality_dimensions": quality_dims,
            # Chunking metadata (P-Q columns, downstream consumption)
            "decomposition_note": decomp_note or "",
            "structure_description": structure or "",
        }
        classifications.append(cls_dict)

    wb.close()
    return classifications


def _parse_quality_dim(text: str, default_name: str) -> dict[str, Any]:
    """Parse a quality dimension cell like ``来源可靠性50%：description...``"""
    # Try "nameXX%：desc" pattern
    m = re.match(r"^(.+?)(\d+)%\s*[:：]\s*(.+)$", text.strip())
    if m:
        name_part = m.group(1).strip()
        weight = int(m.group(2)) / 100.0
        desc = m.group(3).strip()
        return {"name": name_part or default_name, "weight": weight, "description": desc}
    # Fallback: treat whole text as description
    return {"name": default_name, "weight": 0.0, "description": text.strip()}


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def build_rules_content(excel_path: str | Path | None = None) -> dict[str, Any]:
    """Build the complete ``rules_content`` dict from Excel.

    This is the one-time seed data entry point.  Returns a dict ready for
    insertion into ``GovernanceRulesVersion.rules_content``.

    Parameters
    ----------
    excel_path:
        Path to the Excel file.  Defaults to the repo-relative location.

    Returns
    -------
    dict
        Complete rules content with ``schema_version``, ``tag_dimensions``,
        ``classifications``, ``levels``, and ``quality_scoring``.
    """
    return {
        "schema_version": "2.0",
        "tag_dimensions": parse_tag_dimensions(excel_path),
        "classifications": parse_classifications(excel_path),
        "levels": _BUILTIN_LEVELS,
        "quality_scoring": _DEFAULT_QUALITY_SCORING,
        "manual_review_triggers": [],
        "approved_private_model_aliases": [],
    }
