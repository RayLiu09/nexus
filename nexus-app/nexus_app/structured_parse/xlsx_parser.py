"""xlsx → ParsedWorkbook parser (Pipeline B B1.2).

Implements the hard requirements pinned in `docs/pipeline_b_contract_freeze.md
§3.4 / §5.0`:

  - Merged-cell forward-fill (column-direction AND horizontal: every cell
    inside a merged range adopts the origin cell's value).
  - Multi-line cell text is preserved verbatim — no whitespace collapse, no
    flattening of '\\n' / '\\r'.
  - Excel datetimes are normalized to ISO8601 strings; naive datetimes are
    treated as the configured timezone (default Asia/Shanghai).
  - Sheet order and names are preserved (no reorder / no rename).
  - Index columns (序号 / No. / # etc.) are detected by first-row header match
    and dropped from the output, with their column numbers retained in
    ``ParsedSheet.dropped_index_columns`` for audit.
  - Placeholder rows are flagged (``is_placeholder_candidate``) but NOT
    filtered — domain_normalize / quality_validate decides the policy (per
    design §3.4).
  - Every cell carries enough provenance (column, column_letter, merged_range)
    to assemble trace records later, without re-opening the source.

This module deliberately does NOT call MinerU, an LLM, or any governance code
(B0 forbidden changes / decision 8).
"""
from __future__ import annotations

import io
import logging
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import BinaryIO
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from nexus_app.structured_parse.config import (
    DEFAULT_TIMEZONE,
    INDEX_COLUMN_ALIASES,
    PLACEHOLDER_CELL_PATTERNS,
)
from nexus_app.structured_parse.exceptions import (
    CorruptSourceError,
    StructuredParseError,
)
from nexus_app.structured_parse.schemas import (
    CellValue,
    ParsedCell,
    ParsedRow,
    ParsedSheet,
    ParsedWorkbook,
)

logger = logging.getLogger(__name__)

PARSER_VERSION = "xlsx_parser.v1"

# Types accepted by parse_xlsx — we don't accept text streams (xlsx is a
# binary zip container).
XlsxSource = bytes | str | Path | BinaryIO


def parse_xlsx(
    source: XlsxSource,
    *,
    source_filename: str | None = None,
    source_mime_type: str | None = None,
    timezone_name: str = DEFAULT_TIMEZONE,
    index_column_aliases: frozenset[str] | None = None,
    placeholder_patterns: frozenset[str] | None = None,
) -> ParsedWorkbook:
    """Parse xlsx bytes / path / stream into a `ParsedWorkbook`.

    Args:
        source: xlsx bytes, filesystem path, or open binary file-like object.
        source_filename: original filename for the audit trail (recorded on
            the ParsedWorkbook; never trusted for parsing decisions).
        source_mime_type: source mime for the audit trail.
        timezone_name: IANA timezone used to attach to naive datetimes from
            the xlsx (Excel stores datetimes without zone info).
        index_column_aliases: override for the row-number column header set.
        placeholder_patterns: override for the placeholder-row marker set.

    Raises:
        CorruptSourceError: source is not a valid xlsx (bad zip, missing parts).
        StructuredParseError: other unrecoverable parse-time failures.
    """
    tz = _resolve_tz(timezone_name)
    aliases = index_column_aliases if index_column_aliases is not None else INDEX_COLUMN_ALIASES
    placeholders = (
        placeholder_patterns if placeholder_patterns is not None else PLACEHOLDER_CELL_PATTERNS
    )

    wb = _load(source)
    try:
        sheets = [
            _parse_sheet(
                wb[name],
                sheet_index=i,
                tz=tz,
                index_column_aliases=aliases,
                placeholder_patterns=placeholders,
            )
            for i, name in enumerate(wb.sheetnames)
        ]
    finally:
        wb.close()

    return ParsedWorkbook(
        parser_version=PARSER_VERSION,
        parsed_at=datetime.now(tz=timezone.utc),
        source_filename=source_filename,
        source_mime_type=source_mime_type,
        timezone=timezone_name,
        sheets=sheets,
    )


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------


def _load(source: XlsxSource) -> Workbook:
    """Open an xlsx workbook from path / bytes / file-like.

    We deliberately use ``read_only=False`` because ``read_only=True`` does
    not expose the workbook-level ``merged_cells`` metadata we need for
    forward-fill. Memory cost is acceptable for the sample sizes we target.
    """
    handle: BinaryIO | str | Path
    if isinstance(source, bytes):
        handle = io.BytesIO(source)
    else:
        handle = source
    try:
        return load_workbook(handle, data_only=True, read_only=False)
    except (zipfile.BadZipFile, InvalidFileException) as exc:
        raise CorruptSourceError(f"xlsx open failed: {exc}") from exc
    except Exception as exc:  # openpyxl raises various malformed-file errors
        raise StructuredParseError(
            f"xlsx open failed: {type(exc).__name__}: {exc}"
        ) from exc


def _resolve_tz(name: str) -> ZoneInfo:
    try:
        return ZoneInfo(name)
    except ZoneInfoNotFoundError as exc:
        raise StructuredParseError(f"unknown timezone {name!r}: {exc}") from exc


# ---------------------------------------------------------------------------
# Sheet parsing
# ---------------------------------------------------------------------------


def _parse_sheet(
    ws: Worksheet,
    *,
    sheet_index: int,
    tz: ZoneInfo,
    index_column_aliases: frozenset[str],
    placeholder_patterns: frozenset[str],
) -> ParsedSheet:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    if max_row == 0 or max_col == 0:
        return ParsedSheet(
            name=ws.title,
            sheet_index=sheet_index,
            rows=[],
            merged_ranges=[],
            column_count=0,
            row_count=0,
            dropped_index_columns=[],
        )

    # Forward-fill map: every covered (row, col) of every merged range gets
    # mapped to its origin value + range string + origin flag, computed once
    # so per-cell processing stays O(1).
    fill_map = _build_forward_fill_map(ws)
    merged_ranges = sorted(str(r) for r in ws.merged_cells.ranges)
    dropped_index_columns = _detect_index_columns(ws, index_column_aliases)

    normalized_placeholders = {p.strip() for p in placeholder_patterns}

    rows: list[ParsedRow] = []
    for row_idx, row_tuple in enumerate(
        ws.iter_rows(
            min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False
        ),
        start=1,
    ):
        # Placeholder detection MUST happen before column drop — the "……" marker
        # often lives in the index column (e.g. sample 1 row 4: A=……). Dropping
        # the column first would blind us to placeholder semantics.
        any_placeholder = _row_has_placeholder(
            row_tuple, row_idx=row_idx, fill_map=fill_map, patterns=normalized_placeholders
        )

        cells: list[ParsedCell] = []
        all_empty = True
        for col_idx, cell in enumerate(row_tuple, start=1):
            if col_idx in dropped_index_columns:
                continue
            parsed = _build_cell(
                cell, row_idx=row_idx, col_idx=col_idx, fill_map=fill_map, tz=tz
            )
            if parsed.value not in (None, ""):
                all_empty = False
            cells.append(parsed)

        rows.append(
            ParsedRow(
                row_index=row_idx,
                cells=cells,
                is_placeholder_candidate=any_placeholder,
                is_empty=all_empty,
            )
        )

    return ParsedSheet(
        name=ws.title,
        sheet_index=sheet_index,
        rows=rows,
        merged_ranges=merged_ranges,
        column_count=max_col,
        row_count=max_row,
        dropped_index_columns=sorted(dropped_index_columns),
    )


def _row_has_placeholder(
    row_tuple,
    *,
    row_idx: int,
    fill_map: dict[tuple[int, int], tuple[object, str, bool]],
    patterns: set[str],
) -> bool:
    """Return True if ANY cell in the row (including merge-filled and to-be-
    dropped columns) carries a string value matching a placeholder pattern.

    Run before the drop / build loop so the marker can live in any column.
    """
    for col_idx, cell in enumerate(row_tuple, start=1):
        raw = cell.value
        merge_info = fill_map.get((row_idx, col_idx))
        if merge_info is not None and not merge_info[2]:
            # Non-origin cell inside a merge: openpyxl returns None; resolve
            # to the origin value so a placeholder in a merged region is still
            # detected.
            raw = merge_info[0]
        if isinstance(raw, str) and raw.strip() in patterns:
            return True
    return False


def _build_forward_fill_map(
    ws: Worksheet,
) -> dict[tuple[int, int], tuple[object, str, bool]]:
    """Pre-compute origin value + range string + is_origin flag for every cell
    inside any merged range.

    Returns:
        dict mapping (row, col) → (origin_value, range_str, is_origin).
    """
    fill_map: dict[tuple[int, int], tuple[object, str, bool]] = {}
    for merge_range in ws.merged_cells.ranges:
        min_row = merge_range.min_row
        min_col = merge_range.min_col
        range_str = str(merge_range)
        origin_value = ws.cell(row=min_row, column=min_col).value
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            for c in range(merge_range.min_col, merge_range.max_col + 1):
                is_origin = (r, c) == (min_row, min_col)
                fill_map[(r, c)] = (origin_value, range_str, is_origin)
    return fill_map


def _detect_index_columns(ws: Worksheet, aliases: frozenset[str]) -> set[int]:
    """Find columns whose first-row header matches an index-column alias.

    Returns a set of 1-based column indices to drop. Heuristic: only row 1
    is inspected; sheets with multi-row preamble (e.g. title cell at A1) won't
    accidentally drop columns just because the preamble happens to contain an
    alias token in some non-header row.
    """
    detected: set[int] = set()
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row < 1 or max_col < 1:
        return detected
    normalized = {a.strip().casefold() for a in aliases}
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value is None:
            continue
        candidate = str(cell.value).strip().casefold()
        if candidate in normalized:
            detected.add(col)
    return detected


# ---------------------------------------------------------------------------
# Cell construction
# ---------------------------------------------------------------------------


def _build_cell(
    cell,
    *,
    row_idx: int,
    col_idx: int,
    fill_map: dict[tuple[int, int], tuple[object, str, bool]],
    tz: ZoneInfo,
) -> ParsedCell:
    """Convert an openpyxl cell into a ParsedCell.

    Handles merge forward-fill (using the pre-built fill_map so we never
    reach back into openpyxl per-cell), datetime ISO8601 conversion, and
    multi-line flagging.
    """
    column_letter = get_column_letter(col_idx)
    raw_value = cell.value

    is_filled_from_merge = False
    is_merged_origin = False
    merged_range_str: str | None = None

    merge_info = fill_map.get((row_idx, col_idx))
    if merge_info is not None:
        origin_value, range_str, is_origin = merge_info
        merged_range_str = range_str
        if is_origin:
            is_merged_origin = True
            # raw_value already comes from the origin cell (openpyxl stores
            # value on origin only); keep as-is.
        else:
            raw_value = origin_value
            is_filled_from_merge = True

    value, raw_text = _normalize_value(raw_value, tz=tz)
    is_multiline = isinstance(raw_value, str) and ("\n" in raw_value or "\r" in raw_value)

    return ParsedCell(
        column=col_idx,
        column_letter=column_letter,
        value=value,
        raw_text=raw_text,
        is_merged_origin=is_merged_origin,
        is_filled_from_merge=is_filled_from_merge,
        merged_range=merged_range_str,
        is_multiline=is_multiline,
    )


def _normalize_value(raw: object, *, tz: ZoneInfo) -> tuple[CellValue, str | None]:
    """Normalize a raw openpyxl cell value.

    Returns ``(normalized_value, raw_text)``:
      - ``datetime`` → ISO8601 string with explicit timezone; ``raw_text`` = original repr
      - ``str`` / ``int`` / ``float`` / ``bool`` / ``None`` → unchanged value; ``raw_text`` = None
      - Anything else (e.g. openpyxl RichText) → ``str(raw)`` value, ``repr(raw)`` raw_text
    """
    if raw is None:
        return None, None
    if isinstance(raw, datetime):
        aware = raw if raw.tzinfo is not None else raw.replace(tzinfo=tz)
        return aware.isoformat(), str(raw)
    # bool must be checked BEFORE int because in Python `bool` is a subclass of int.
    if isinstance(raw, bool):
        return raw, None
    if isinstance(raw, (str, int, float)):
        return raw, None
    return str(raw), repr(raw)


__all__ = [
    "PARSER_VERSION",
    "parse_xlsx",
]
