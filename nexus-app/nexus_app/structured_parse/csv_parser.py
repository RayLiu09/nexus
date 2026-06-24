"""csv → ParsedWorkbook parser (Pipeline B B1.4).

Single-sheet representation: CSV is flat by definition, so the resulting
ParsedWorkbook always has exactly one ParsedSheet whose name is derived from
the caller-supplied `sheet_name` (default `"csv"`).

Implements the hard requirements from `docs/pipeline_b_contract_freeze.md
§3.4 / §5.0`:

  - Multi-line text inside quoted fields is preserved verbatim
    (Python's csv module handles RFC 4180 escaping).
  - Index columns (序号 / No. / # etc.) are detected by first-row header match
    and dropped with their column numbers retained in
    `ParsedSheet.dropped_index_columns`.
  - Placeholder rows are flagged but NOT filtered.
  - Cell text is preserved verbatim; CSV has no type system, so all cell
    values are strings (or None for blank cells per RFC 4180 quoting).

Encoding strategy (no auto-detect by default):
  - Caller may pass an explicit `encoding`.
  - Otherwise try UTF-8 first (with BOM stripped); fall back to GB18030 for
    Chinese-language exports from Excel; on both failures raise
    CorruptSourceError.

Datetime semantics:
  - CSV has no datetime type; all fields stay as strings.
  - The `timezone_name` argument is kept for API symmetry with `parse_xlsx`
    and recorded on the ParsedWorkbook for the audit trail.
"""
from __future__ import annotations

import codecs
import csv
import io
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import BinaryIO, TextIO
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl.utils import get_column_letter

from nexus_app.structured_parse.config import (
    DEFAULT_TIMEZONE,
    INDEX_COLUMN_ALIASES,
    PLACEHOLDER_CELL_PATTERNS,
)
from nexus_app.structured_parse.exceptions import (
    CorruptSourceError,
    StructuredParseError,
)
from nexus_app.structured_parse.schemas import (
    ParsedCell,
    ParsedRow,
    ParsedSheet,
    ParsedWorkbook,
)

logger = logging.getLogger(__name__)

PARSER_VERSION = "csv_parser.v1"

CsvSource = bytes | str | Path | BinaryIO | TextIO

# Encodings tried in order when caller omits `encoding`. GB18030 is the
# superset of GBK / GB2312 commonly emitted by Excel for Chinese data.
_DEFAULT_ENCODING_FALLBACKS: tuple[str, ...] = ("utf-8-sig", "utf-8", "gb18030")


def parse_csv(
    source: CsvSource,
    *,
    source_filename: str | None = None,
    source_mime_type: str | None = "text/csv",
    timezone_name: str = DEFAULT_TIMEZONE,
    index_column_aliases: frozenset[str] | None = None,
    placeholder_patterns: frozenset[str] | None = None,
    encoding: str | None = None,
    delimiter: str = ",",
    sheet_name: str = "csv",
) -> ParsedWorkbook:
    """Parse CSV bytes / str / path / file-like into a ParsedWorkbook.

    Args:
        source: CSV bytes, text, filesystem path, or open file-like.
        encoding: Force a specific encoding. When None, auto-detect from
            (utf-8 with BOM, utf-8, gb18030).
        delimiter: CSV field delimiter (default `,`).
        sheet_name: Logical sheet name recorded on the single ParsedSheet.

    Raises:
        CorruptSourceError: bytes not decodable as any candidate encoding.
        StructuredParseError: unknown timezone or other unrecoverable error.
    """
    # Validate timezone early — fail before doing any I/O. CSV itself has no
    # datetime cells, but the timezone is recorded on the ParsedWorkbook for
    # downstream consumers and audit symmetry with parse_xlsx.
    try:
        ZoneInfo(timezone_name)
    except ZoneInfoNotFoundError as exc:
        raise StructuredParseError(f"unknown timezone {timezone_name!r}: {exc}") from exc

    text = _decode(source, encoding=encoding)

    aliases = index_column_aliases if index_column_aliases is not None else INDEX_COLUMN_ALIASES
    placeholders = (
        placeholder_patterns if placeholder_patterns is not None else PLACEHOLDER_CELL_PATTERNS
    )

    rows = _parse_rows(text, delimiter=delimiter)
    column_count = max((len(r) for r in rows), default=0)
    dropped_index_columns = _detect_index_columns(rows, aliases)
    normalized_placeholders = {p.strip() for p in placeholders}

    parsed_rows: list[ParsedRow] = []
    for row_idx, raw_row in enumerate(rows, start=1):
        any_placeholder = _row_has_placeholder(raw_row, normalized_placeholders)
        cells: list[ParsedCell] = []
        all_empty = True
        for col_idx in range(1, column_count + 1):
            if col_idx in dropped_index_columns:
                continue
            value = raw_row[col_idx - 1] if col_idx - 1 < len(raw_row) else None
            normalized_value = value if value else None  # treat empty string as None
            cells.append(
                ParsedCell(
                    column=col_idx,
                    column_letter=get_column_letter(col_idx),
                    value=normalized_value,
                    raw_text=None,
                    is_merged_origin=False,
                    is_filled_from_merge=False,
                    merged_range=None,
                    is_multiline=isinstance(normalized_value, str)
                    and ("\n" in normalized_value or "\r" in normalized_value),
                )
            )
            if normalized_value not in (None, ""):
                all_empty = False
        parsed_rows.append(
            ParsedRow(
                row_index=row_idx,
                cells=cells,
                is_placeholder_candidate=any_placeholder,
                is_empty=all_empty,
            )
        )

    sheet = ParsedSheet(
        name=sheet_name,
        sheet_index=0,
        rows=parsed_rows,
        merged_ranges=[],  # CSV has no merged cells
        column_count=column_count,
        row_count=len(parsed_rows),
        dropped_index_columns=sorted(dropped_index_columns),
    )

    return ParsedWorkbook(
        parser_version=PARSER_VERSION,
        parsed_at=datetime.now(tz=timezone.utc),
        source_filename=source_filename,
        source_mime_type=source_mime_type,
        timezone=timezone_name,
        sheets=[sheet],
    )


# ---------------------------------------------------------------------------
# Decoding
# ---------------------------------------------------------------------------


def _decode(source: CsvSource, *, encoding: str | None) -> str:
    """Resolve a CsvSource into a decoded string.

    Order:
      1. bytes / path / binary-mode file → read bytes, try encoding(s) until success
      2. str → already-decoded, BOM stripped if present
      3. text-mode file → read as-is
    """
    raw_bytes: bytes | None = None
    text_value: str | None = None

    if isinstance(source, str) and not _looks_like_path(source):
        text_value = source
    elif isinstance(source, (str, Path)):
        path = Path(source)
        try:
            raw_bytes = path.read_bytes()
        except OSError as exc:
            raise CorruptSourceError(f"csv path read failed: {exc}") from exc
    elif isinstance(source, bytes):
        raw_bytes = source
    elif hasattr(source, "read"):
        # File-like — could be binary or text mode
        try:
            content = source.read()
        except Exception as exc:
            raise CorruptSourceError(f"csv stream read failed: {exc}") from exc
        if isinstance(content, bytes):
            raw_bytes = content
        elif isinstance(content, str):
            text_value = content
        else:
            raise StructuredParseError(
                f"csv source stream returned unexpected type: {type(content).__name__}"
            )
    else:
        raise StructuredParseError(
            f"csv source must be bytes, str, Path, or file-like; got {type(source).__name__}"
        )

    if text_value is not None:
        # Strip BOM if present
        if text_value.startswith(codecs.BOM_UTF8.decode("utf-8")):
            text_value = text_value[1:]
        return text_value

    assert raw_bytes is not None
    if encoding is not None:
        candidates = (encoding,)
    else:
        candidates = _DEFAULT_ENCODING_FALLBACKS

    last_err: UnicodeDecodeError | None = None
    for enc in candidates:
        try:
            decoded = raw_bytes.decode(enc)
            return decoded
        except UnicodeDecodeError as exc:
            last_err = exc
            continue
    raise CorruptSourceError(
        f"csv decode failed with all candidate encodings ({candidates}): {last_err}"
    )


def _looks_like_path(value: str) -> bool:
    """Heuristic: treat a short, no-newline, no-comma str as a path.

    CSV content is overwhelmingly multi-line or comma-laden; a bare path like
    `data.csv` lacks both. This lets callers pass either CSV text or a path
    string without explicit Path() wrapping.

    Empty / whitespace-only strings are treated as CSV text (not paths) so
    `parse_csv("")` produces an empty workbook rather than trying to read
    the current working directory.
    """
    if not value or value.isspace():
        return False
    if "\n" in value or "\r" in value:
        return False
    if "," in value:
        return False
    if len(value) > 4096:
        return False
    try:
        return Path(value).is_file()
    except OSError:
        return False


# ---------------------------------------------------------------------------
# Row parsing
# ---------------------------------------------------------------------------


def _parse_rows(text: str, *, delimiter: str) -> list[list[str]]:
    """RFC 4180 row split via Python's csv module.

    Handles:
      - CRLF / LF line endings
      - Quoted fields with embedded delimiters / newlines
      - Doubled quotes inside quoted fields
    """
    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
    try:
        return [list(row) for row in reader]
    except csv.Error as exc:
        raise CorruptSourceError(f"csv parse failed: {exc}") from exc


def _detect_index_columns(rows: list[list[str]], aliases: frozenset[str]) -> set[int]:
    """Mirror `xlsx_parser._detect_index_columns`: row-1 header match only.

    Returns 1-based column indices to drop.
    """
    detected: set[int] = set()
    if not rows:
        return detected
    normalized = {a.strip().casefold() for a in aliases}
    header = rows[0]
    for col_idx, value in enumerate(header, start=1):
        if value is None:
            continue
        candidate = str(value).strip().casefold()
        if candidate in normalized:
            detected.add(col_idx)
    return detected


def _row_has_placeholder(row: list[str], patterns: set[str]) -> bool:
    for cell in row:
        if isinstance(cell, str) and cell.strip() in patterns:
            return True
    return False


__all__ = [
    "PARSER_VERSION",
    "parse_csv",
]
