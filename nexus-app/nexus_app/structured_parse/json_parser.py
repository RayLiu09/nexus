"""json → ParsedWorkbook parser (Pipeline B B1.4).

Scope (P0):
  - Top-level array of objects:  ``[{...}, {...}]``
  - Wrapper object with a record-array field at one of
    ``record_path_candidates`` (default: ``records / items / data / rows``):
    ``{"records": [{...}, ...], "meta": {...}}``
  - Single object treated as a single-row dataset: ``{"k": "v"}``

Column derivation:
  - Header set = union of all record keys, preserving first-occurrence order.
  - Missing keys in a record produce a None cell in that column.
  - Nested arrays / objects inside a record are JSON-serialised to a string
    (stored in `value`) and the original repr is kept in `raw_text` for audit.

Out of scope (left for future parsers):
  - JSON Lines (one JSON value per line) — would need its own helper.
  - Crawler payloads with bespoke envelope shapes — `_load_record_payload`
    (worker) keeps handling those until B2 / B4 introduce profile-driven
    dispatch.

This parser is NOT yet wired into `execute_job()` (per B1.4 scope decision):
the existing JSON path through `_load_record_payload` is left untouched so
crawler / database / webhook ingestion contracts stay stable. B2 may later
route file_upload+JSON through `parse_json` once profile_detect can decide
between business-object and table-shaped JSON.
"""
from __future__ import annotations

import codecs
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, BinaryIO, TextIO
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl.utils import get_column_letter

from nexus_app.structured_parse.config import (
    DEFAULT_TIMEZONE,
    INDEX_COLUMN_ALIASES,
    PLACEHOLDER_CELL_PATTERNS,
)
from nexus_app.structured_parse.exceptions import (
    CorruptSourceError,
    EmptySourceError,
    StructuredParseError,
)
from nexus_app.structured_parse.schemas import (
    CellValue,
    ParsedCell,
    ParsedRow,
    ParsedSheet,
    ParsedWorkbook,
)

logger = logging.getLogger(__name__)

PARSER_VERSION = "json_parser.v1"

JsonSource = bytes | str | Path | BinaryIO | TextIO

_DEFAULT_RECORD_PATH_CANDIDATES: tuple[str, ...] = ("records", "items", "data", "rows")


def parse_json(
    source: JsonSource,
    *,
    source_filename: str | None = None,
    source_mime_type: str | None = "application/json",
    timezone_name: str = DEFAULT_TIMEZONE,
    index_column_aliases: frozenset[str] | None = None,
    placeholder_patterns: frozenset[str] | None = None,
    encoding: str = "utf-8",
    record_path_candidates: tuple[str, ...] = _DEFAULT_RECORD_PATH_CANDIDATES,
    sheet_name: str = "json",
) -> ParsedWorkbook:
    """Parse JSON bytes / str / path / file-like into a ParsedWorkbook.

    Args:
        record_path_candidates: When the top-level JSON is an object, the
            parser walks these keys in order and picks the first whose value
            is a list of objects. Override to support custom wrappers.

    Raises:
        CorruptSourceError: source bytes are not valid JSON.
        EmptySourceError: source decoded but contained no records / properties.
        StructuredParseError: unknown timezone or unsupported source type.
    """
    try:
        ZoneInfo(timezone_name)
    except ZoneInfoNotFoundError as exc:
        raise StructuredParseError(f"unknown timezone {timezone_name!r}: {exc}") from exc

    text = _decode(source, encoding=encoding)
    parsed = _load_json(text)
    records = _extract_records(parsed, record_path_candidates=record_path_candidates)

    if not records:
        raise EmptySourceError("json source contained no records")

    aliases = index_column_aliases if index_column_aliases is not None else INDEX_COLUMN_ALIASES
    placeholders = (
        placeholder_patterns if placeholder_patterns is not None else PLACEHOLDER_CELL_PATTERNS
    )
    normalized_placeholders = {p.strip() for p in placeholders}

    header = _derive_header(records)
    dropped_index_columns = _detect_index_columns(header, aliases)

    rows: list[ParsedRow] = [_build_header_row(header, dropped_index_columns)]
    for data_idx, record in enumerate(records, start=2):  # row 1 is the header
        any_placeholder = _record_has_placeholder(record, normalized_placeholders)
        cells: list[ParsedCell] = []
        all_empty = True
        for col_idx, key in enumerate(header, start=1):
            if col_idx in dropped_index_columns:
                continue
            raw_value = record.get(key) if isinstance(record, dict) else None
            value, raw_text = _normalize_value(raw_value)
            is_multiline = isinstance(value, str) and ("\n" in value or "\r" in value)
            cells.append(
                ParsedCell(
                    column=col_idx,
                    column_letter=get_column_letter(col_idx),
                    value=value,
                    raw_text=raw_text,
                    is_merged_origin=False,
                    is_filled_from_merge=False,
                    merged_range=None,
                    is_multiline=is_multiline,
                )
            )
            if value not in (None, ""):
                all_empty = False
        rows.append(
            ParsedRow(
                row_index=data_idx,
                cells=cells,
                is_placeholder_candidate=any_placeholder,
                is_empty=all_empty,
            )
        )

    sheet = ParsedSheet(
        name=sheet_name,
        sheet_index=0,
        rows=rows,
        merged_ranges=[],
        column_count=len(header),
        row_count=len(rows),
        dropped_index_columns=sorted(dropped_index_columns),
    )

    return ParsedWorkbook(
        parser_version=PARSER_VERSION,
        parsed_at=datetime.now(tz=timezone.utc),
        source_filename=source_filename,
        source_mime_type=source_mime_type,
        timezone=timezone_name,
        sheets=[sheet],
    )


# ---------------------------------------------------------------------------
# Decoding & loading
# ---------------------------------------------------------------------------


def _decode(source: JsonSource, *, encoding: str) -> str:
    """Resolve a JsonSource into a decoded string. JSON is always UTF-8 per
    RFC 8259 unless the caller specifies otherwise; we accept the caller's
    override only if utf-8 fails."""
    raw_bytes: bytes | None = None
    text_value: str | None = None

    if isinstance(source, str) and not _looks_like_path(source):
        text_value = source
    elif isinstance(source, (str, Path)):
        path = Path(source)
        try:
            raw_bytes = path.read_bytes()
        except OSError as exc:
            raise CorruptSourceError(f"json path read failed: {exc}") from exc
    elif isinstance(source, bytes):
        raw_bytes = source
    elif hasattr(source, "read"):
        try:
            content = source.read()
        except Exception as exc:
            raise CorruptSourceError(f"json stream read failed: {exc}") from exc
        if isinstance(content, bytes):
            raw_bytes = content
        elif isinstance(content, str):
            text_value = content
        else:
            raise StructuredParseError(
                f"json source stream returned unexpected type: {type(content).__name__}"
            )
    else:
        raise StructuredParseError(
            f"json source must be bytes, str, Path, or file-like; got {type(source).__name__}"
        )

    if text_value is not None:
        if text_value.startswith(codecs.BOM_UTF8.decode("utf-8")):
            text_value = text_value[1:]
        return text_value

    assert raw_bytes is not None
    # Strip a UTF-8 BOM transparently so a caller-supplied utf-8 / utf-8-sig
    # mix doesn't trip Python's `json.loads` (which rejects BOM since 3.6+).
    if raw_bytes.startswith(codecs.BOM_UTF8):
        raw_bytes = raw_bytes[len(codecs.BOM_UTF8):]
    try:
        return raw_bytes.decode(encoding)
    except UnicodeDecodeError as exc:
        raise CorruptSourceError(f"json decode failed with {encoding!r}: {exc}") from exc


def _looks_like_path(value: str) -> bool:
    """Treat a short bare path-like string as a path; treat JSON content (which
    almost always contains `{` / `[` / quotes) as text."""
    if not value or value.isspace():
        return False
    if "\n" in value or "\r" in value:
        return False
    if any(ch in value for ch in "{}[]\""):
        return False
    if len(value) > 4096:
        return False
    try:
        return Path(value).is_file()
    except OSError:
        return False


def _load_json(text: str) -> Any:
    try:
        return json.loads(text)
    except json.JSONDecodeError as exc:
        raise CorruptSourceError(f"json parse failed: {exc}") from exc


# ---------------------------------------------------------------------------
# Record extraction
# ---------------------------------------------------------------------------


def _extract_records(
    parsed: Any, *, record_path_candidates: tuple[str, ...]
) -> list[dict[str, Any]]:
    """Normalise the parsed JSON into a list of dict-shaped records.

    Three accepted shapes:
      1. Top-level list → each element must be a dict.
      2. Top-level dict containing a record-array under a known key.
      3. Top-level dict treated as a single-row record (when no record-array
         key is found and there's at least one property).

    Raises:
        CorruptSourceError when shape is unsupported (e.g. list of scalars,
        nested arrays without an array of dicts).
    """
    if isinstance(parsed, list):
        if not parsed:
            return []
        # Require homogeneous dict elements; abort if any element is non-dict.
        for i, item in enumerate(parsed):
            if not isinstance(item, dict):
                raise CorruptSourceError(
                    f"json top-level array element at index {i} is not an object "
                    f"(got {type(item).__name__}); structured_parse requires "
                    f"an array of objects"
                )
        return parsed

    if isinstance(parsed, dict):
        for key in record_path_candidates:
            candidate = parsed.get(key)
            if isinstance(candidate, list) and candidate:
                if all(isinstance(item, dict) for item in candidate):
                    return candidate
                # Fall through — wrapper found but not array-of-dicts; try other keys.
        # No matching wrapper; treat the dict itself as a single-row record.
        if parsed:
            return [parsed]
        return []

    raise CorruptSourceError(
        f"json top-level value must be an object or an array of objects; "
        f"got {type(parsed).__name__}"
    )


def _derive_header(records: list[dict[str, Any]]) -> list[str]:
    """Union of record keys in first-occurrence order."""
    seen: dict[str, None] = {}  # dict preserves insertion order
    for record in records:
        for key in record.keys():
            if key not in seen:
                seen[key] = None
    return list(seen.keys())


def _detect_index_columns(header: list[str], aliases: frozenset[str]) -> set[int]:
    detected: set[int] = set()
    if not header:
        return detected
    normalized = {a.strip().casefold() for a in aliases}
    for col_idx, name in enumerate(header, start=1):
        if name is None:
            continue
        if str(name).strip().casefold() in normalized:
            detected.add(col_idx)
    return detected


def _record_has_placeholder(record: dict[str, Any], patterns: set[str]) -> bool:
    for value in record.values():
        if isinstance(value, str) and value.strip() in patterns:
            return True
    return False


def _build_header_row(header: list[str], dropped: set[int]) -> ParsedRow:
    """Synthesise a header row (row_index=1) so downstream consumers can treat
    JSON parsers and xlsx/csv parsers uniformly (first row is always header).

    Dropped index columns are excluded here too — keeping them in the header
    while removing them from data rows would create a column-count mismatch
    that downstream profile_detect / domain_normalize would have to special-case.
    """
    cells = [
        ParsedCell(
            column=i,
            column_letter=get_column_letter(i),
            value=name,
            raw_text=None,
        )
        for i, name in enumerate(header, start=1)
        if i not in dropped
    ]
    return ParsedRow(row_index=1, cells=cells, is_empty=not cells)


def _normalize_value(raw: Any) -> tuple[CellValue, str | None]:
    """Normalise a raw JSON value for ParsedCell.

    JSON natively yields: str / int / float / bool / None / list / dict.
    The first five fit ParsedCell.value directly. Nested list / dict are
    serialised to a JSON string in `value`, with `raw_text` holding the
    original repr for audit (so a future profile_detect can detect that the
    column carries structured sub-data).
    """
    if raw is None:
        return None, None
    if isinstance(raw, bool):  # bool BEFORE int (bool is subclass of int)
        return raw, None
    if isinstance(raw, (str, int, float)):
        return raw, None
    if isinstance(raw, (list, dict)):
        serialised = json.dumps(raw, ensure_ascii=False)
        return serialised, repr(raw)
    return str(raw), repr(raw)


__all__ = [
    "PARSER_VERSION",
    "parse_json",
]
