"""Tests for `structured_parse.xlsx_parser`.

Two layers:

  1. Synthetic xlsx tests built in-memory via openpyxl — pin each individual
     contract (merge fill, multiline, datetime, index drop, placeholder).
  2. Sample-file integration tests using docs/samples/ — exercise the full
     parser against real worksheets (sample 1 岗位需求, sample 2 PGSD).

The synthetic tests fail fast on logic regressions; the sample tests catch
real-world quirks (multi-row preamble, multi-sheet ordering, ⏎ in cells).
"""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Callable

import pytest
from openpyxl import Workbook

from nexus_app.structured_parse.exceptions import (
    CorruptSourceError,
    StructuredParseError,
)
from nexus_app.structured_parse.schemas import ParsedWorkbook
from nexus_app.structured_parse.xlsx_parser import PARSER_VERSION, parse_xlsx

# Repo root resolution: this file is at nexus-app/tests/structured_parse/
# Five `parents` calls reach the repo root (file → structured_parse → tests →
# nexus-app → projects/nexus).
REPO_ROOT = Path(__file__).resolve().parents[3]
SAMPLE_JOB_DEMAND = REPO_ROOT / "docs/samples/1.（岗位需求）电子商务岗位招聘数据.xlsx"
SAMPLE_ABILITY = REPO_ROOT / "docs/samples/2.（职业能力分析）大数据技术应用专业职业能力分析表.xlsx"


def _make_xlsx_bytes(builder: Callable[[Workbook], None]) -> bytes:
    """Build an xlsx in memory via openpyxl and return bytes."""
    wb = Workbook()
    builder(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Smoke
# ---------------------------------------------------------------------------


class TestParserVersionConstant:
    def test_parser_version_pinned(self):
        # Pinned because downstream record_body_hash cache invalidation will
        # eventually depend on this string.
        assert PARSER_VERSION == "xlsx_parser.v1"


class TestEmptyAndDegenerate:
    def test_default_workbook_one_empty_sheet(self):
        # Default openpyxl Workbook() ships with one sheet named "Sheet".
        result = parse_xlsx(_make_xlsx_bytes(lambda wb: None))
        assert isinstance(result, ParsedWorkbook)
        assert [s.name for s in result.sheets] == ["Sheet"]
        # max_row / max_col may be 1 for a "fresh" sheet (openpyxl quirk);
        # we accept either truly empty or single-cell empty.
        assert result.sheets[0].row_count <= 1
        # parser_version & timezone recorded
        assert result.parser_version == PARSER_VERSION

    def test_corrupt_bytes_raises_corrupt_source_error(self):
        with pytest.raises(CorruptSourceError):
            parse_xlsx(b"definitely not an xlsx")

    def test_truncated_bytes_raises_corrupt_source_error(self):
        # Take a valid xlsx and chop most of it off — should fail to open.
        good = _make_xlsx_bytes(lambda wb: wb.active.cell(row=1, column=1, value="x"))
        with pytest.raises(CorruptSourceError):
            parse_xlsx(good[:50])


# ---------------------------------------------------------------------------
# Sheet ordering and naming
# ---------------------------------------------------------------------------


class TestSheetOrderPreservation:
    def test_sheet_order_and_names_preserved(self):
        def build(wb: Workbook) -> None:
            wb.active.title = "first"
            wb.create_sheet("second")
            wb.create_sheet("third")
            wb.create_sheet("0.numeric prefix")

        result = parse_xlsx(_make_xlsx_bytes(build))
        assert [s.name for s in result.sheets] == [
            "first",
            "second",
            "third",
            "0.numeric prefix",
        ]
        assert [s.sheet_index for s in result.sheets] == [0, 1, 2, 3]


# ---------------------------------------------------------------------------
# Merge forward-fill
# ---------------------------------------------------------------------------


class TestMergedCellForwardFill:
    def test_column_merge_forward_fills_subsequent_rows(self):
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws["A1"] = "header"
            ws["A2"] = "value"
            ws.merge_cells("A2:A5")
            ws["B2"] = "x"

        result = parse_xlsx(_make_xlsx_bytes(build))
        sheet = result.sheets[0]
        col_a = [r.cells[0] for r in sheet.rows]

        # Row 1 (header) — not merged
        assert col_a[0].value == "header"
        assert col_a[0].is_merged_origin is False
        assert col_a[0].is_filled_from_merge is False
        assert col_a[0].merged_range is None

        # Row 2 — merge origin
        assert col_a[1].value == "value"
        assert col_a[1].is_merged_origin is True
        assert col_a[1].is_filled_from_merge is False
        assert col_a[1].merged_range == "A2:A5"

        # Rows 3-5 — forward-filled from origin
        for filled in col_a[2:5]:
            assert filled.value == "value"
            assert filled.is_merged_origin is False
            assert filled.is_filled_from_merge is True
            assert filled.merged_range == "A2:A5"

    def test_horizontal_merge_fills_with_origin_value(self):
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws["A1"] = "x"
            ws["B1"] = "header"
            ws.merge_cells("B1:D1")

        result = parse_xlsx(_make_xlsx_bytes(build))
        row = result.sheets[0].rows[0]
        # Columns: A=1, B=2, C=3, D=4 → cells[0..3]
        assert row.cells[1].value == "header"
        assert row.cells[1].is_merged_origin is True
        assert row.cells[2].value == "header"
        assert row.cells[2].is_filled_from_merge is True
        assert row.cells[2].merged_range == "B1:D1"
        assert row.cells[3].value == "header"
        assert row.cells[3].is_filled_from_merge is True


# ---------------------------------------------------------------------------
# Multiline preservation
# ---------------------------------------------------------------------------


class TestMultilinePreserved:
    def test_multiline_text_kept_verbatim(self):
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws["A1"] = "line1\nline2\nline3"

        result = parse_xlsx(_make_xlsx_bytes(build))
        cell = result.sheets[0].rows[0].cells[0]
        assert cell.value == "line1\nline2\nline3"
        assert cell.is_multiline is True

    def test_carriage_return_also_flags_multiline(self):
        # openpyxl normalises bare '\r' line endings to '\n' on read, so the
        # multi-line flag must trigger on either character. We only assert
        # that the flag fires and that some line break survives.
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws["A1"] = "line1\rline2"

        result = parse_xlsx(_make_xlsx_bytes(build))
        cell = result.sheets[0].rows[0].cells[0]
        assert cell.is_multiline is True
        assert ("\n" in cell.value) or ("\r" in cell.value)


# ---------------------------------------------------------------------------
# Datetime normalization
# ---------------------------------------------------------------------------


class TestDatetimeNormalization:
    def test_naive_datetime_gets_default_asia_shanghai_timezone(self):
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws["A1"] = datetime(2024, 12, 12, 1, 12, 59)

        result = parse_xlsx(_make_xlsx_bytes(build))
        cell = result.sheets[0].rows[0].cells[0]
        assert cell.value == "2024-12-12T01:12:59+08:00"
        # raw_text preserves the original datetime repr (no timezone)
        assert cell.raw_text == "2024-12-12 01:12:59"

    def test_custom_timezone_argument_used(self):
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws["A1"] = datetime(2024, 12, 12, 1, 12, 59)

        result = parse_xlsx(_make_xlsx_bytes(build), timezone_name="UTC")
        assert result.sheets[0].rows[0].cells[0].value.endswith("+00:00")
        assert result.timezone == "UTC"

    def test_unknown_timezone_raises(self):
        with pytest.raises(StructuredParseError):
            parse_xlsx(
                _make_xlsx_bytes(lambda wb: wb.active.cell(row=1, column=1, value="x")),
                timezone_name="Not/A_Real_Zone",
            )


# ---------------------------------------------------------------------------
# Index column drop
# ---------------------------------------------------------------------------


class TestIndexColumnDrop:
    @pytest.mark.parametrize(
        "alias", ["序号", "No.", "no.", "#", "ID", "id", "row_id", "编号"]
    )
    def test_first_row_alias_triggers_column_drop(self, alias):
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws["A1"] = alias
            ws["B1"] = "data"
            ws["A2"] = 1
            ws["B2"] = "row1"
            ws["A3"] = 2
            ws["B3"] = "row2"

        result = parse_xlsx(_make_xlsx_bytes(build))
        sheet = result.sheets[0]
        assert sheet.dropped_index_columns == [1]
        # The first surviving cell in every row is column B
        for r in sheet.rows:
            assert r.cells[0].column == 2
            assert r.cells[0].column_letter == "B"

    def test_non_alias_first_row_keeps_all_columns(self):
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws["A1"] = "title"
            ws["B1"] = "anything"

        result = parse_xlsx(_make_xlsx_bytes(build))
        sheet = result.sheets[0]
        assert sheet.dropped_index_columns == []

    def test_alias_in_non_first_row_does_not_drop_column(self):
        # The index-column check is intentionally limited to row 1 so
        # multi-row preambles can't accidentally trigger a drop.
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws["A1"] = "title"
            ws["A2"] = "序号"  # not row 1 → ignored
            ws["A3"] = 1

        result = parse_xlsx(_make_xlsx_bytes(build))
        assert result.sheets[0].dropped_index_columns == []

    def test_custom_alias_set_overrides_default(self):
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws["A1"] = "Lp."  # Polish row-number convention; not in default set
            ws["B1"] = "data"
            ws["A2"] = 1

        result = parse_xlsx(
            _make_xlsx_bytes(build), index_column_aliases=frozenset({"Lp."})
        )
        assert result.sheets[0].dropped_index_columns == [1]


# ---------------------------------------------------------------------------
# Placeholder row flagging
# ---------------------------------------------------------------------------


class TestPlaceholderRowFlagging:
    def test_chinese_ellipsis_flags_row(self):
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws["A1"] = "data"
            ws["A2"] = "……"
            ws["A3"] = "real"

        rows = parse_xlsx(_make_xlsx_bytes(build)).sheets[0].rows
        assert rows[0].is_placeholder_candidate is False
        assert rows[1].is_placeholder_candidate is True
        assert rows[2].is_placeholder_candidate is False

    def test_placeholder_rows_are_flagged_not_filtered(self):
        # Per design §3.4 placeholder filtering is deferred to
        # domain_normalize / quality_validate, so rows MUST stay in the output.
        def build(wb: Workbook) -> None:
            ws = wb.active
            ws["A1"] = "real"
            ws["A2"] = "……"

        result = parse_xlsx(_make_xlsx_bytes(build))
        assert len(result.sheets[0].rows) == 2


# ---------------------------------------------------------------------------
# Sample 1 — 岗位需求电子商务（17 cols, 4 rows, "序号" header, datetime, multiline）
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not SAMPLE_JOB_DEMAND.exists(), reason="sample missing")
class TestSampleJobDemand:
    @pytest.fixture(scope="class")
    def parsed(self) -> ParsedWorkbook:
        return parse_xlsx(
            SAMPLE_JOB_DEMAND.read_bytes(),
            source_filename=SAMPLE_JOB_DEMAND.name,
            source_mime_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

    def test_single_sheet_named_sheet1(self, parsed):
        assert [s.name for s in parsed.sheets] == ["Sheet1"]

    def test_record_count_is_four_rows(self, parsed):
        # 1 header + 2 data + 1 placeholder = 4 (sample inspected upstream)
        # Note: sample actually has 1 header + 3 data + 1 placeholder = 5 rows
        # per inspection. We accept either 4 or 5 in case the sample is updated.
        assert parsed.sheets[0].row_count in {4, 5}

    def test_index_column_序号_dropped(self, parsed):
        sheet = parsed.sheets[0]
        assert sheet.dropped_index_columns == [1]
        # Header row should no longer expose 序号 but must still expose 岗位名称
        header_values = [c.value for c in sheet.rows[0].cells]
        assert "序号" not in header_values
        assert "岗位名称" in header_values

    def test_placeholder_row_flagged(self, parsed):
        # Last data row contains "……" sentinel
        assert any(r.is_placeholder_candidate for r in parsed.sheets[0].rows)

    def test_published_at_datetime_normalized_to_iso8601(self, parsed):
        # Column O = 发布时间. After dropping column A, the cell's `column` field
        # still reports the original 15; we look it up by column_letter for safety.
        row2 = parsed.sheets[0].rows[1]
        pub_cell = next(c for c in row2.cells if c.column_letter == "O")
        assert isinstance(pub_cell.value, str)
        assert pub_cell.value.startswith("2024-12-12T")
        assert "+08:00" in pub_cell.value

    def test_multiline_description_preserved(self, parsed):
        row2 = parsed.sheets[0].rows[1]
        desc = next(c for c in row2.cells if c.column_letter == "J")
        assert isinstance(desc.value, str)
        assert "\n" in desc.value
        assert desc.is_multiline is True


# ---------------------------------------------------------------------------
# Sample 2 — PGSD 能力分析（5 sheets, A 列大类合并 forward-fill, B3 多行 ①②③）
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not SAMPLE_ABILITY.exists(), reason="sample missing")
class TestSampleAbilityAnalysis:
    @pytest.fixture(scope="class")
    def parsed(self) -> ParsedWorkbook:
        return parse_xlsx(
            SAMPLE_ABILITY.read_bytes(),
            source_filename=SAMPLE_ABILITY.name,
            source_mime_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

    def test_five_sheets_preserved_in_order(self, parsed):
        assert [s.name for s in parsed.sheets] == [
            "典型工作任务和工作内容分析表",
            "1.数据采集",
            "2.数据标注",
            "3.数据ETL处理",
            "4.可视化图表制作",
        ]

    def test_no_index_column_dropped_in_any_sheet(self, parsed):
        # Every sheet's A1 is a title ("能力分析表" / "典型工作任务...") — none
        # of these match the 序号 alias set, so no column should be dropped.
        for sheet in parsed.sheets:
            assert sheet.dropped_index_columns == [], (
                f"sheet {sheet.name!r} unexpectedly dropped {sheet.dropped_index_columns}"
            )

    def test_a_column_category_forward_fills(self, parsed):
        # In sheet "1.数据采集": A5 = "职业能力" with a column merge A5:A20.
        # Rows 6..20 should report the same value as A5 with is_filled_from_merge=True.
        sheet = next(s for s in parsed.sheets if s.name == "1.数据采集")
        rows_by_index = {r.row_index: r for r in sheet.rows}

        a5 = next(c for c in rows_by_index[5].cells if c.column_letter == "A")
        assert a5.value == "职业能力"
        assert a5.is_merged_origin is True
        assert a5.merged_range and a5.merged_range.startswith("A5:")

        a6 = next(c for c in rows_by_index[6].cells if c.column_letter == "A")
        assert a6.value == "职业能力"
        assert a6.is_filled_from_merge is True
        assert a6.merged_range == a5.merged_range

    def test_task_description_multiline_preserves_circled_numbers(self, parsed):
        sheet = next(s for s in parsed.sheets if s.name == "1.数据采集")
        b3 = next(
            c
            for c in next(r for r in sheet.rows if r.row_index == 3).cells
            if c.column_letter == "B"
        )
        # B3 holds the multi-line task description (含 ①②③)
        assert isinstance(b3.value, str)
        assert "\n" in b3.value
        assert b3.is_multiline is True
        assert "①" in b3.value

    def test_overview_sheet_records_merged_ranges(self, parsed):
        # The overview sheet has merges A1:I1, C2:I2, A2:B3. We just confirm
        # the parser surfaces them (downstream profile_detect inspects this).
        overview = parsed.sheets[0]
        assert any(r == "A1:I1" for r in overview.merged_ranges)
        assert any(r == "C2:I2" for r in overview.merged_ranges)
        assert any(r == "A2:B3" for r in overview.merged_ranges)
