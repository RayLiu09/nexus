"""Tests for the Pipeline B B2.3 worker integration of `profile_detect`.

Three layers:

  1. Unit: `_run_profile_detect()` — detect call + RECORD_PROFILE_DETECTED
     audit + fallback when the detector itself raises (defensive).
  2. Unit: `_maybe_park_in_review_required()` — candidate / generic / low-
     confidence trigger version_status transition + audit; canonical
     high-confidence stays in PROCESSING.
  3. Integration: `execute_job()` end-to-end — xlsx high-confidence stays
     PROCESSING with `metadata_summary.profile` populated; synthetic
     candidate xlsx transitions to REVIEW_REQUIRED with both audits.
"""
from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import select

from nexus_app import models, services
from nexus_app.config import get_settings
from nexus_app.enums import (
    AssetVersionStatus,
    AuditEventType,
    DataSourceType,
    IngestBatchStatus,
    JobStatus,
    JobType,
    RawObjectStatus,
)
from nexus_app.mineru import FakeMinerUAdapter
from nexus_app.pipeline.payload_schema import JOB_PAYLOAD_SCHEMA_VERSION
from nexus_app.profile_detect import (
    DEFAULT_AUTO_ADMIT_THRESHOLD,
    DETECTOR_VERSION,
    ProfileDetectResult,
    ProfileEvidence,
)
from nexus_app.schemas import DataSourceCreate
from nexus_app.storage import InMemoryObjectStorage
from nexus_app.worker.runner import (
    _maybe_park_in_review_required,
    _run_profile_detect,
    execute_job,
)

REPO_ROOT = Path(__file__).resolve().parents[2]
SAMPLE_JOB_DEMAND = REPO_ROOT / "docs/samples/1.（岗位需求）电子商务岗位招聘数据.xlsx"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------------------------------------------------------------------
# Synthetic xlsx helpers — keep the integration tests self-contained for the
# candidate / negative paths the docs/samples/ files don't cover.
# ---------------------------------------------------------------------------


def _make_job_demand_xlsx_bytes() -> bytes:
    """Full recruiting-header workbook — should detect at high confidence."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["岗位名称", "城市", "公司名称", "薪资", "学历要求", "岗位描述", "发布时间"])
    ws.append(["平面设计师", "上海", "ACME", "10k-15k", "本科", "...", "2025-01"])
    ws.append(["开发", "北京", "Foo", "20k-30k", "硕士", "...", "2025-02"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_candidate_xlsx_bytes() -> bytes:
    """Only one required header — detector downgrades to candidate."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["岗位名称", "filler1", "filler2"])
    ws.append(["dev", "x", "y"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_generic_xlsx_bytes() -> bytes:
    """No recognised signals → generic_table_dataset fallback."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["alpha", "beta", "gamma"])
    ws.append(["1", "2", "3"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _seed_xlsx_raw(
    session, storage: InMemoryObjectStorage, *, xlsx_bytes: bytes, source_code: str
):
    """Construct a queued xlsx record-pipeline job for direct execute_job invocation."""
    source = services.create_data_source(
        session,
        DataSourceCreate(code=source_code, name=source_code, source_type="file_upload"),
    )
    batch = models.IngestBatch(
        data_source_id=source.id,
        idempotency_key=f"{source_code}-key",
        source_type=DataSourceType.FILE_UPLOAD,
        status=IngestBatchStatus.RAW_PERSISTED,
    )
    session.add(batch)
    session.flush()
    key = f"raw/{source_code}.xlsx"
    stored = storage.put_bytes(key, xlsx_bytes, XLSX_MIME)
    raw = models.RawObject(
        data_source_id=source.id,
        batch_id=batch.id,
        source_type=DataSourceType.FILE_UPLOAD,
        object_uri=stored.object_uri,
        checksum=stored.checksum,
        mime_type=XLSX_MIME,
        size_bytes=stored.size_bytes,
        status=RawObjectStatus.RAW_PERSISTED,
        metadata_summary={"filename": f"{source_code}.xlsx"},
    )
    session.add(raw)
    session.flush()
    job = models.Job(
        job_type=JobType.INGEST_PROCESS,
        status=JobStatus.RUNNING,
        ingest_batch_id=batch.id,
        raw_object_id=raw.id,
        idempotency_key=f"{source_code}-key",
        payload={"pipeline_type": "record"},
        payload_schema_version=JOB_PAYLOAD_SCHEMA_VERSION,
        trace_id=f"trace-{source_code}",
    )
    session.add(job)
    session.commit()
    return job, raw, batch


def _audits(session, event_type: AuditEventType) -> list[models.AuditLog]:
    return list(
        session.scalars(
            select(models.AuditLog).where(models.AuditLog.event_type == event_type)
        )
    )


# ---------------------------------------------------------------------------
# Unit: _run_profile_detect — defensive contract: never raises
# ---------------------------------------------------------------------------


class TestRunProfileDetect:
    def test_returns_detector_result_for_valid_workbook(self, session):
        # Minimal in-memory raw_payload mimicking ParsedWorkbook.model_dump
        raw_payload = {
            "parser_version": "xlsx_parser.v1",
            "parsed_at": "2026-06-25T00:00:00+00:00",
            "timezone": "Asia/Shanghai",
            "source_filename": "x.xlsx",
            "source_mime_type": XLSX_MIME,
            "sheets": [
                {
                    "name": "Sheet1",
                    "sheet_index": 0,
                    "rows": [
                        {
                            "row_index": 1,
                            "cells": [
                                {"column": 1, "column_letter": "A", "value": "岗位名称"},
                                {"column": 2, "column_letter": "B", "value": "城市"},
                                {"column": 3, "column_letter": "C", "value": "公司名称"},
                                {"column": 4, "column_letter": "D", "value": "薪资"},
                            ],
                        },
                    ],
                    "merged_ranges": [],
                    "column_count": 4,
                    "row_count": 1,
                    "dropped_index_columns": [],
                },
            ],
        }
        # Minimal job / raw_object stand-ins — the helper only reads their IDs.
        job = models.Job(
            job_type=JobType.INGEST_PROCESS,
            status=JobStatus.RUNNING,
            payload={"pipeline_type": "record"},
            payload_schema_version=JOB_PAYLOAD_SCHEMA_VERSION,
        )
        session.add(job)
        session.flush()
        raw = models.RawObject(
            data_source_id="ds-x",
            batch_id="batch-x",
            source_type=DataSourceType.FILE_UPLOAD,
            object_uri="s3://bucket/k",
            checksum="abc",
            mime_type=XLSX_MIME,
            size_bytes=1,
            status=RawObjectStatus.RAW_PERSISTED,
            metadata_summary={"filename": "x.xlsx"},
        )
        session.add(raw)
        session.flush()

        result = _run_profile_detect(job, raw_payload, raw, session, "trace-x")
        assert isinstance(result, ProfileDetectResult)
        # 3 required headers + 1 optional → solid hit on job_demand
        assert result.record_type in {"job_demand_dataset", "job_demand_dataset_candidate"}

        audits = _audits(session, AuditEventType.RECORD_PROFILE_DETECTED)
        assert len(audits) == 1
        assert audits[0].summary["record_type"] == result.record_type
        assert audits[0].summary["confidence"] == result.confidence
        assert audits[0].summary["detector_version"] == DETECTOR_VERSION
        assert audits[0].target_id == raw.id

    def test_falls_back_to_generic_table_on_unexpected_payload(self, session):
        # Garbage payload — ParsedWorkbook.model_validate will reject it;
        # the helper must NOT raise (worker still needs a result for the
        # downstream version-state branch).
        job = models.Job(
            job_type=JobType.INGEST_PROCESS,
            status=JobStatus.RUNNING,
            payload={"pipeline_type": "record"},
            payload_schema_version=JOB_PAYLOAD_SCHEMA_VERSION,
        )
        session.add(job)
        session.flush()
        raw = models.RawObject(
            data_source_id="ds-x",
            batch_id="batch-x",
            source_type=DataSourceType.FILE_UPLOAD,
            object_uri="s3://bucket/k",
            checksum="abc",
            mime_type=XLSX_MIME,
            size_bytes=1,
            status=RawObjectStatus.RAW_PERSISTED,
        )
        session.add(raw)
        session.flush()

        garbage_payload = {"not": "a workbook"}
        result = _run_profile_detect(job, garbage_payload, raw, session, "trace-y")
        assert result.record_type == "generic_table_dataset"
        assert result.confidence == 0.0


# ---------------------------------------------------------------------------
# Unit: _maybe_park_in_review_required — the trigger matrix
# ---------------------------------------------------------------------------


def _profile(record_type: str, confidence: float, *, analysis_model: str | None = None) -> ProfileDetectResult:
    return ProfileDetectResult(
        record_type=record_type,  # type: ignore[arg-type]
        domain="occupation",
        domain_profile="job_demand.v1" if "job_demand" in record_type else "ability_analysis.pgsd.v1",
        analysis_model=analysis_model,
        detector_version=DETECTOR_VERSION,
        confidence=confidence,
        evidence=ProfileEvidence(),
    )


_attach_counter = 0


def _attach_version(session):
    """Build a minimal AssetVersion in PROCESSING state for the trigger tests."""
    global _attach_counter
    _attach_counter += 1
    n = _attach_counter
    asset = models.Asset(
        data_source_id=f"ds-vp-{n}",
        source_object_key=f"key-vp-{n}",
        title=f"vp-{n}",
        asset_kind="record",
    )
    session.add(asset)
    session.flush()
    raw = models.RawObject(
        data_source_id=f"ds-vp-{n}",
        batch_id=f"batch-vp-{n}",
        source_type=DataSourceType.FILE_UPLOAD,
        object_uri=f"s3://b/k-{n}",
        checksum=f"cs-{n}",
        mime_type=XLSX_MIME,
        size_bytes=1,
        status=RawObjectStatus.RAW_PERSISTED,
    )
    session.add(raw)
    session.flush()
    version = models.AssetVersion(
        asset_id=asset.id,
        raw_object_id=raw.id,
        version_no=1,
        version_status=AssetVersionStatus.PROCESSING,
        source_checksum=f"vp-chk-{n}",
    )
    session.add(version)
    session.flush()
    return version, raw


class TestMaybeParkInReviewRequired:
    def test_high_confidence_canonical_stays_processing(self, session):
        version, raw = _attach_version(session)
        profile = _profile("job_demand_dataset", 0.95)
        transitioned = _maybe_park_in_review_required(
            profile, version, raw, session, "trace-h", "job-h"
        )
        assert transitioned is False
        assert version.version_status == AssetVersionStatus.PROCESSING
        assert _audits(session, AuditEventType.VERSION_STATUS_CHANGED) == []
        assert _audits(session, AuditEventType.RECORD_PROFILE_REVIEW_REQUIRED) == []

    @pytest.mark.parametrize(
        "record_type",
        [
            "job_demand_dataset_candidate",
            "occupational_ability_analysis_candidate",
            "generic_table_dataset",
        ],
    )
    def test_candidate_or_generic_transitions_to_review_required(self, session, record_type):
        version, raw = _attach_version(session)
        # Use a high confidence so the trigger is the record_type, not the
        # threshold — proves both branches of the OR cover their case.
        profile = _profile(record_type, 0.99,
                            analysis_model="PGSD" if "ability" in record_type else None)
        transitioned = _maybe_park_in_review_required(
            profile, version, raw, session, "trace-c", "job-c"
        )
        assert transitioned is True
        assert version.version_status == AssetVersionStatus.REVIEW_REQUIRED

        version_audits = _audits(session, AuditEventType.VERSION_STATUS_CHANGED)
        assert len(version_audits) == 1
        assert version_audits[0].summary["previous_status"] == "processing"
        assert version_audits[0].summary["current_status"] == "review_required"
        assert version_audits[0].summary["reason"] == "profile_detect_candidate_or_low_confidence"

        review_audits = _audits(session, AuditEventType.RECORD_PROFILE_REVIEW_REQUIRED)
        assert len(review_audits) == 1
        assert review_audits[0].summary["record_type"] == record_type

    def test_low_confidence_below_threshold_also_triggers(self, session):
        # Canonical record_type with confidence < threshold — defence-in-depth
        # branch (in normal flow the dispatcher would have downgraded to
        # candidate, but a custom-threshold caller could land here).
        version, raw = _attach_version(session)
        profile = _profile("job_demand_dataset", DEFAULT_AUTO_ADMIT_THRESHOLD - 0.01)
        transitioned = _maybe_park_in_review_required(
            profile, version, raw, session, "trace-l", "job-l"
        )
        assert transitioned is True
        assert version.version_status == AssetVersionStatus.REVIEW_REQUIRED

    def test_idempotent_no_duplicate_version_status_audit(self, session):
        # If the version is already REVIEW_REQUIRED, we should NOT emit a
        # second VERSION_STATUS_CHANGED audit (governance_decision must not
        # double-fire on retries). RECORD_PROFILE_REVIEW_REQUIRED still
        # fires every time as it's the canonical "profile reviewed" event.
        version, raw = _attach_version(session)
        version.version_status = AssetVersionStatus.REVIEW_REQUIRED
        session.flush()

        profile = _profile("generic_table_dataset", 0.05)
        _maybe_park_in_review_required(
            profile, version, raw, session, "trace-i", "job-i"
        )
        assert _audits(session, AuditEventType.VERSION_STATUS_CHANGED) == []
        assert len(_audits(session, AuditEventType.RECORD_PROFILE_REVIEW_REQUIRED)) == 1


# ---------------------------------------------------------------------------
# Integration: execute_job end-to-end (xlsx record pipeline)
# ---------------------------------------------------------------------------


class TestExecuteJobIntegration:
    def test_full_recruiting_xlsx_writes_profile_into_normalized_ref(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_xlsx_raw(
            session, storage,
            xlsx_bytes=_make_job_demand_xlsx_bytes(),
            source_code="b23-high",
        )

        execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())

        # job succeeded
        session.refresh(job)
        assert job.status == JobStatus.SUCCEEDED, f"job failed: {job.failure_reason}"

        # NormalizedAssetRef carries profile in metadata_summary (PG-side mirror)
        refs = list(session.scalars(select(models.NormalizedAssetRef)))
        assert len(refs) == 1
        ref = refs[0]
        ms_profile = ref.metadata_summary.get("profile")
        assert ms_profile is not None, "expected profile mirrored into metadata_summary"
        assert ms_profile["record_type"] == "job_demand_dataset"
        assert ms_profile["domain_profile"] == "job_demand.v1"
        assert ms_profile["confidence"] >= DEFAULT_AUTO_ADMIT_THRESHOLD

        # MinIO-side normalized_payload also has the top-level profile field
        payload = json.loads(
            storage.get_bytes(ref.object_uri.split("/", 3)[-1]).decode("utf-8")
        )
        assert payload["profile"]["record_type"] == "job_demand_dataset"
        # `record_type` at the payload root is the detected one, not the
        # raw_object metadata fallback ("generic").
        assert payload["record_type"] == "job_demand_dataset"

        # Audits: DETECTED once, REVIEW_REQUIRED never
        assert len(_audits(session, AuditEventType.RECORD_PROFILE_DETECTED)) == 1
        assert _audits(session, AuditEventType.RECORD_PROFILE_REVIEW_REQUIRED) == []

        # version stays in PROCESSING — governance_decision (skipped here)
        # would otherwise drive it forward
        versions = list(session.scalars(select(models.AssetVersion)))
        assert versions[0].version_status == AssetVersionStatus.PROCESSING

    def test_candidate_xlsx_parks_version_in_review_required(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_xlsx_raw(
            session, storage,
            xlsx_bytes=_make_candidate_xlsx_bytes(),
            source_code="b23-cand",
        )

        execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())

        session.refresh(job)
        assert job.status == JobStatus.SUCCEEDED

        refs = list(session.scalars(select(models.NormalizedAssetRef)))
        ms_profile = refs[0].metadata_summary["profile"]
        assert ms_profile["record_type"] == "job_demand_dataset_candidate"

        # version transitioned to review_required
        versions = list(session.scalars(select(models.AssetVersion)))
        assert versions[0].version_status == AssetVersionStatus.REVIEW_REQUIRED

        # Both detection and review-required audits fired
        assert len(_audits(session, AuditEventType.RECORD_PROFILE_DETECTED)) == 1
        assert len(_audits(session, AuditEventType.RECORD_PROFILE_REVIEW_REQUIRED)) == 1
        # And the version-state transition audit fired exactly once
        vsc_audits = [
            a for a in _audits(session, AuditEventType.VERSION_STATUS_CHANGED)
            if a.summary.get("reason") == "profile_detect_candidate_or_low_confidence"
        ]
        assert len(vsc_audits) == 1

    def test_generic_xlsx_also_parks_in_review_required(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_xlsx_raw(
            session, storage,
            xlsx_bytes=_make_generic_xlsx_bytes(),
            source_code="b23-gen",
        )

        execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())

        session.refresh(job)
        refs = list(session.scalars(select(models.NormalizedAssetRef)))
        ms_profile = refs[0].metadata_summary["profile"]
        assert ms_profile["record_type"] == "generic_table_dataset"

        versions = list(session.scalars(select(models.AssetVersion)))
        assert versions[0].version_status == AssetVersionStatus.REVIEW_REQUIRED


# ---------------------------------------------------------------------------
# Integration: real sample 1 (job_demand) — clears the auto-admit threshold
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not SAMPLE_JOB_DEMAND.exists(), reason="sample missing")
class TestSampleJobDemandE2E:
    def test_sample1_detected_with_high_confidence(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_xlsx_raw(
            session, storage,
            xlsx_bytes=SAMPLE_JOB_DEMAND.read_bytes(),
            source_code="b23-sample1",
        )

        execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())

        session.refresh(job)
        assert job.status == JobStatus.SUCCEEDED

        refs = list(session.scalars(select(models.NormalizedAssetRef)))
        ms_profile = refs[0].metadata_summary["profile"]
        assert ms_profile["record_type"] == "job_demand_dataset"
        assert ms_profile["confidence"] >= 0.90

        # PROCESSING (not REVIEW_REQUIRED) since canonical record_type at
        # full confidence
        versions = list(session.scalars(select(models.AssetVersion)))
        assert versions[0].version_status == AssetVersionStatus.PROCESSING
