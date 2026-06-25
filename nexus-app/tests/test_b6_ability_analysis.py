"""B6 ability_analysis writer / seed / dispatcher / E2E tests.

Layers covered here (writer-side; API tests live in
`nexus-api/tests/test_b6_ability_analysis_api.py`):

  1. PGSD seed: verifies the `ability_analysis_profile` seed row matches the
     contract (model_code, schema_version, four categories, code patterns).
  2. Writer:  field mapping, upsert, quality_flags, blocking vs non-blocking
     rejection rules, relation emission, audit events.
  3. Dispatcher: integration through `dispatch_domain_normalize` reading
     `record_body` from in-memory MinIO.
  4. End-to-end (sample 2): canonical record_body built from the real xlsx
     → dispatcher → writer → tables populated.

Sample-2 E2E uses a small canonical record_body builder (kept test-local)
rather than relying on a future B3 transformer — the B6 contract only
specifies that the writer consumes the contract-shape `record_body`, not
how it is built. A real B3+ transformer landing later is decoupled from
this test.
"""
from __future__ import annotations

import io
import json
import uuid
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from nexus_app import models
from nexus_app.domain_normalize import dispatch_domain_normalize
from nexus_app.domain_normalize.ability_analysis_writer import write as ability_writer_write
from nexus_app.enums import (
    AuditEventType,
    NormalizedAssetRefStatus,
    NormalizedType,
)
from nexus_app.storage import InMemoryObjectStorage

REPO_ROOT = Path(__file__).resolve().parents[2]
_SAMPLE_REL = "docs/samples/2.（职业能力分析）大数据技术应用专业职业能力分析表.xlsx"


def _resolve_sample_xlsx() -> Path:
    """`docs/samples/` is gitignored, so when this test file is checked
    out in a git worktree the sample is absent. Fall back to the main
    repo's checkout under `~/projects/nexus/docs/samples/` if found, so
    the sample-E2E tests still run when the worktree is sibling to the
    main repo.
    """
    primary = REPO_ROOT / _SAMPLE_REL
    if primary.exists():
        return primary
    fallback = Path("/home/bjbodao/projects/nexus") / _SAMPLE_REL
    return fallback if fallback.exists() else primary


SAMPLE_ABILITY_XLSX = _resolve_sample_xlsx()


# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------


PGSD_CATEGORY_SCHEMA = [
    {"code": "P", "name": "职业能力", "alias": ["职业技能"]},
    {"code": "G", "name": "通用能力"},
    {"code": "S", "name": "社会能力"},
    {"code": "D", "name": "发展能力"},
]


PGSD_CODE_PATTERN = {
    "P": {
        "regex": r"^P-\d+\.\d+\.\d+$",
        "segments": 3,
        "requires_work_content": True,
    },
    "G": {
        "regex": r"^G-\d+\.\d+$",
        "segments": 2,
        "requires_work_content": False,
    },
    "S": {
        "regex": r"^S-\d+\.\d+$",
        "segments": 2,
        "requires_work_content": False,
    },
    "D": {
        "regex": r"^D-\d+\.\d+$",
        "segments": 2,
        "requires_work_content": False,
    },
}


def _seed_pgsd_profile(session: Session) -> models.AbilityAnalysisProfile:
    """SQLite tests don't run alembic migrations (conftest uses
    Base.metadata.create_all), so the PGSD seed has to be created manually
    here. Mirrors the data written by alembic 0044.
    """
    existing = session.scalar(
        select(models.AbilityAnalysisProfile).where(
            models.AbilityAnalysisProfile.model_code == "PGSD",
            models.AbilityAnalysisProfile.schema_version == "ability_analysis.pgsd.v1",
        )
    )
    if existing is not None:
        return existing
    profile = models.AbilityAnalysisProfile(
        model_code="PGSD",
        model_name="职业能力分析 PGSD 模型",
        schema_version="ability_analysis.pgsd.v1",
        category_schema=PGSD_CATEGORY_SCHEMA,
        code_pattern=PGSD_CODE_PATTERN,
        relation_schema={},
        detector_rules={},
        is_active=True,
        is_builtin=True,
        initialized_by="system_seed",
        initialized_at=datetime.now(timezone.utc),
    )
    session.add(profile)
    session.flush()
    return profile


def _make_ref_with_payload(
    session: Session,
    storage: InMemoryObjectStorage,
    *,
    record_body: dict,
    domain_profile: str = "ability_analysis.pgsd.v1",
    trace_id: str = "trace-b6",
) -> models.NormalizedAssetRef:
    """Build an `asset → asset_version → normalized_asset_ref` triple and
    store the canonical payload in `storage` so the dispatcher can load it.
    """
    from nexus_app.enums import (
        AssetKind,
        AssetVersionStatus,
        DataSourceType,
        IngestBatchStatus,
        RawObjectStatus,
    )
    suffix = uuid.uuid4().hex[:8]
    ds = models.DataSource(
        code=f"b6-ds-{suffix}",
        name=f"b6-ds-{suffix}",
        source_type=DataSourceType.FILE_UPLOAD,
    )
    session.add(ds)
    session.flush()
    batch = models.IngestBatch(
        data_source_id=ds.id,
        idempotency_key=f"b6-key-{suffix}",
        source_type=DataSourceType.FILE_UPLOAD,
        status=IngestBatchStatus.RAW_PERSISTED,
    )
    session.add(batch)
    session.flush()
    raw = models.RawObject(
        data_source_id=ds.id,
        batch_id=batch.id,
        source_type=DataSourceType.FILE_UPLOAD,
        object_uri=f"s3://nexus-test-objects/raw/{suffix}.xlsx",
        checksum=f"raw-cs-{suffix}",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        size_bytes=10,
        status=RawObjectStatus.RAW_PERSISTED,
    )
    session.add(raw)
    session.flush()
    asset = models.Asset(
        data_source_id=ds.id,
        source_object_key=f"asset/{suffix}",
        title=f"b6-asset-{suffix}",
        asset_kind=AssetKind.RECORD,
    )
    session.add(asset)
    session.flush()
    version = models.AssetVersion(
        asset_id=asset.id,
        raw_object_id=raw.id,
        version_no=1,
        version_status=AssetVersionStatus.PROCESSING,
        source_checksum=f"src-cs-{suffix}",
    )
    session.add(version)
    session.flush()

    payload = {
        "schema_version": "normalized-record.v2",
        "domain_profile": domain_profile,
        "record_body": record_body,
        "metadata": {"domain_profile": domain_profile},
    }
    payload_key = f"normalized/{suffix}.json"
    storage.put_bytes(payload_key, json.dumps(payload).encode("utf-8"), "application/json")
    ref = models.NormalizedAssetRef(
        version_id=version.id,
        normalized_type=NormalizedType.RECORD,
        object_uri=f"s3://nexus-test-objects/{payload_key}",
        schema_version="normalized-record.v2",
        checksum=f"ref-cs-{suffix}",
        status=NormalizedAssetRefStatus.GENERATED,
        metadata_summary={"domain_profile": domain_profile},
        lineage={"trace_id": trace_id},
    )
    session.add(ref)
    session.flush()
    return ref


def _minimal_record_body(**overrides) -> dict:
    """Tiny canonical record_body — one task, one work_content, one P ability."""
    base = {
        "analysis": {
            "major_name": "大数据技术应用",
            "major_direction": None,
            "analysis_model": "PGSD",
            "task_count": 1,
            "work_content_count": 1,
            "ability_item_count": 1,
        },
        "tasks": [
            {
                "task_code": "1",
                "task_name": "数据采集",
                "task_description": "①清洗 ②上传 ③校验",
                "task_description_structured": None,
                "display_order": 1,
                "trace": {"sheet": "1.数据采集", "row": 3},
                "work_contents": [
                    {
                        "content_code": "1.1",
                        "content_name": "日志系统数据采集",
                        "abilities": [
                            {
                                "ability_code": "P-1.1.1",
                                "ability_major_category_code": "P",
                                "ability_content": "能够采集日志",
                                "trace": {"sheet": "1.数据采集", "row": 5},
                            }
                        ],
                    }
                ],
                "general_abilities": {},
            }
        ],
    }
    base.update(overrides)
    return base


def _audits(session: Session, event_type: AuditEventType) -> list[models.AuditLog]:
    return list(
        session.scalars(
            select(models.AuditLog).where(models.AuditLog.event_type == event_type)
        )
    )


# ===========================================================================
# 1. PGSD profile seed
# ===========================================================================


class TestPgsdProfileSeed:
    def test_seed_helper_inserts_one_profile(self, session):
        _seed_pgsd_profile(session)
        rows = list(session.scalars(select(models.AbilityAnalysisProfile)))
        assert len(rows) == 1
        assert rows[0].model_code == "PGSD"
        assert rows[0].schema_version == "ability_analysis.pgsd.v1"
        assert rows[0].is_active is True
        assert rows[0].is_builtin is True

    def test_seed_helper_is_idempotent(self, session):
        _seed_pgsd_profile(session)
        _seed_pgsd_profile(session)
        rows = list(session.scalars(select(models.AbilityAnalysisProfile)))
        assert len(rows) == 1

    def test_category_schema_has_pgsd_four_categories(self, session):
        profile = _seed_pgsd_profile(session)
        codes = sorted(entry["code"] for entry in profile.category_schema)
        assert codes == ["D", "G", "P", "S"]

    def test_p_category_alias_includes_legacy_label(self, session):
        profile = _seed_pgsd_profile(session)
        p_entry = next(e for e in profile.category_schema if e["code"] == "P")
        assert p_entry["name"] == "职业能力"
        assert "职业技能" in p_entry.get("alias", [])

    def test_p_pattern_requires_work_content(self, session):
        profile = _seed_pgsd_profile(session)
        assert profile.code_pattern["P"]["requires_work_content"] is True
        assert profile.code_pattern["P"]["segments"] == 3

    @pytest.mark.parametrize("cat", ["G", "S", "D"])
    def test_gsd_patterns_do_not_require_work_content(self, session, cat):
        profile = _seed_pgsd_profile(session)
        assert profile.code_pattern[cat]["requires_work_content"] is False
        assert profile.code_pattern[cat]["segments"] == 2

    def test_unique_constraint_on_model_schema(self, session):
        from sqlalchemy.exc import IntegrityError
        _seed_pgsd_profile(session)
        duplicate = models.AbilityAnalysisProfile(
            model_code="PGSD",
            model_name="dup",
            schema_version="ability_analysis.pgsd.v1",
            category_schema=[],
            code_pattern={},
        )
        session.add(duplicate)
        with pytest.raises(IntegrityError):
            session.flush()
        session.rollback()


# ===========================================================================
# 2. Writer — field mappings, blocking flags, idempotency
# ===========================================================================


class TestWriterHappyPath:
    @pytest.fixture
    def state(self, session):
        storage = InMemoryObjectStorage()
        _seed_pgsd_profile(session)
        ref = _make_ref_with_payload(
            session, storage, record_body=_minimal_record_body()
        )
        result = ability_writer_write(session, ref, _minimal_record_body())
        return result, ref, storage

    def test_returns_analysis_id(self, state):
        result, _, _ = state
        assert result.analysis_id is not None
        assert result.domain_profile == "ability_analysis.pgsd.v1"

    def test_analysis_row_created(self, session, state):
        result, ref, _ = state
        analysis = session.get(models.OccupationalAbilityAnalysis, result.analysis_id)
        assert analysis is not None
        assert analysis.normalized_ref_id == ref.id
        assert analysis.analysis_model == "PGSD"
        assert analysis.major_name == "大数据技术应用"
        assert analysis.schema_version == "ability_analysis.pgsd.v1"

    def test_task_row_created(self, session, state):
        result, _, _ = state
        tasks = list(
            session.scalars(
                select(models.OccupationalWorkTask).where(
                    models.OccupationalWorkTask.analysis_id == result.analysis_id
                )
            )
        )
        assert len(tasks) == 1
        assert tasks[0].task_code == "1"
        assert tasks[0].task_name == "数据采集"
        # B6 must always write {} here (B5 owns task_description_structured)
        assert tasks[0].task_description_structured == {}

    def test_work_content_row_created(self, session, state):
        result, _, _ = state
        wcs = list(
            session.scalars(
                select(models.OccupationalWorkContent).where(
                    models.OccupationalWorkContent.analysis_id == result.analysis_id
                )
            )
        )
        assert len(wcs) == 1
        assert wcs[0].content_code == "1.1"
        assert wcs[0].content_name == "日志系统数据采集"

    def test_ability_row_created(self, session, state):
        result, _, _ = state
        items = list(
            session.scalars(
                select(models.OccupationalAbilityItem).where(
                    models.OccupationalAbilityItem.analysis_id == result.analysis_id
                )
            )
        )
        assert len(items) == 1
        item = items[0]
        assert item.ability_code == "P-1.1.1"
        assert item.ability_major_category_code == "P"
        # writer must reverse-lookup the canonical name from category_schema
        assert item.ability_major_category_name == "职业能力"
        # P ability must hang off a work_content
        assert item.work_content_id is not None
        # writer extracts sequence by stripping the `<category>-` prefix
        assert item.ability_sequence == "1.1.1"
        # B6 does not set confidence (B5 does)
        assert item.confidence is None

    def test_counts_rolled_up(self, session, state):
        result, _, _ = state
        analysis = session.get(models.OccupationalAbilityAnalysis, result.analysis_id)
        assert analysis.task_count == 1
        assert analysis.work_content_count == 1
        assert analysis.ability_item_count == 1

    def test_relations_emitted(self, session, state):
        result, _, _ = state
        relations = list(
            session.scalars(
                select(models.OccupationalAbilityRelation).where(
                    models.OccupationalAbilityRelation.analysis_id == result.analysis_id
                )
            )
        )
        rel_types = sorted({r.relation_type for r in relations})
        assert rel_types == ["TASK_HAS_WORK_CONTENT", "WORK_CONTENT_REQUIRES_ABILITY"]
        # exactly one of each for a single ability under a single work_content
        assert len(relations) == 2

    def test_audit_events_written(self, session, state):
        persisted = _audits(session, AuditEventType.ABILITY_ANALYSIS_PERSISTED)
        items_persisted = _audits(session, AuditEventType.ABILITY_ITEMS_PERSISTED)
        rejected = _audits(session, AuditEventType.ABILITY_ITEMS_REJECTED)
        assert len(persisted) == 1
        assert len(items_persisted) == 1
        # no rejections in happy path
        assert rejected == []
        assert persisted[0].summary["model_code"] == "PGSD"
        assert persisted[0].summary["task_count"] == 1

    def test_source_dataset_table_not_written_by_default(self, session, state):
        # P0 default: B6 does NOT populate ability_analysis_source_dataset.
        links = list(session.scalars(select(models.AbilityAnalysisSourceDataset)))
        assert links == []


class TestWriterGSDAbilities:
    def _record_body_with_gsd(self) -> dict:
        return {
            "analysis": {
                "major_name": "test_major",
                "analysis_model": "PGSD",
                "task_count": 1,
                "work_content_count": 1,
                "ability_item_count": 4,
            },
            "tasks": [
                {
                    "task_code": "1",
                    "task_name": "T1",
                    "task_description": None,
                    "display_order": 1,
                    "trace": {},
                    "work_contents": [
                        {
                            "content_code": "1.1",
                            "content_name": "C1",
                            "abilities": [
                                {
                                    "ability_code": "P-1.1.1",
                                    "ability_major_category_code": "P",
                                    "ability_content": "p ability",
                                }
                            ],
                        }
                    ],
                    "general_abilities": {
                        "G": [
                            {
                                "ability_code": "G-1.1",
                                "ability_content": "通用能力1",
                            }
                        ],
                        "S": [
                            {
                                "ability_code": "S-1.1",
                                "ability_content": "社会能力1",
                            }
                        ],
                        "D": [
                            {
                                "ability_code": "D-1.1",
                                "ability_content": "发展能力1",
                            }
                        ],
                    },
                }
            ],
        }

    def test_gsd_abilities_persisted_without_work_content(self, session):
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        body = self._record_body_with_gsd()
        ref = _make_ref_with_payload(session, storage, record_body=body)
        result = ability_writer_write(session, ref, body)

        items = list(
            session.scalars(
                select(models.OccupationalAbilityItem).where(
                    models.OccupationalAbilityItem.analysis_id == result.analysis_id
                )
            )
        )
        assert len(items) == 4  # 1 P + 1 G + 1 S + 1 D
        by_code = {it.ability_code: it for it in items}
        # P type hangs off a work_content
        assert by_code["P-1.1.1"].work_content_id is not None
        # G/S/D types have NULL work_content_id (requires_work_content=False)
        assert by_code["G-1.1"].work_content_id is None
        assert by_code["S-1.1"].work_content_id is None
        assert by_code["D-1.1"].work_content_id is None
        assert by_code["G-1.1"].ability_major_category_name == "通用能力"
        assert by_code["S-1.1"].ability_major_category_name == "社会能力"
        assert by_code["D-1.1"].ability_major_category_name == "发展能力"

    def test_gsd_abilities_do_not_emit_work_content_relation(self, session):
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        body = self._record_body_with_gsd()
        ref = _make_ref_with_payload(session, storage, record_body=body)
        result = ability_writer_write(session, ref, body)
        relations = list(
            session.scalars(
                select(models.OccupationalAbilityRelation).where(
                    models.OccupationalAbilityRelation.analysis_id == result.analysis_id
                )
            )
        )
        # 1 TASK_HAS_WORK_CONTENT + 1 WORK_CONTENT_REQUIRES_ABILITY (for P only).
        rel_counts: dict[str, int] = {}
        for r in relations:
            rel_counts[r.relation_type] = rel_counts.get(r.relation_type, 0) + 1
        assert rel_counts == {
            "TASK_HAS_WORK_CONTENT": 1,
            "WORK_CONTENT_REQUIRES_ABILITY": 1,
        }


class TestWriterQualityFlags:
    def test_ability_code_pattern_mismatch_is_non_blocking(self, session):
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        body = _minimal_record_body()
        # P-1.1 is two-segment which violates the P (three-segment) regex.
        body["tasks"][0]["work_contents"][0]["abilities"][0]["ability_code"] = "P-1.1"
        ref = _make_ref_with_payload(session, storage, record_body=body)
        result = ability_writer_write(session, ref, body)

        items = list(session.scalars(select(models.OccupationalAbilityItem)))
        assert len(items) == 1, "non-blocking flag must still let row land"
        assert items[0].quality_flags.get("ability_code_pattern_mismatch") is True
        analysis = session.get(models.OccupationalAbilityAnalysis, result.analysis_id)
        assert analysis.quality_summary.get("ability_code_pattern_mismatch") == 1

    def test_ability_category_unknown_is_blocking(self, session):
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        body = _minimal_record_body()
        body["tasks"][0]["work_contents"][0]["abilities"].append(
            {
                "ability_code": "Z-9.9.9",
                "ability_major_category_code": "Z",   # not in PGSD
                "ability_content": "alien category",
            }
        )
        ref = _make_ref_with_payload(session, storage, record_body=body)
        result = ability_writer_write(session, ref, body)
        items = list(session.scalars(select(models.OccupationalAbilityItem)))
        # Only the original P-1.1.1 row landed; Z-9.9.9 was rejected.
        assert len(items) == 1
        analysis = session.get(models.OccupationalAbilityAnalysis, result.analysis_id)
        assert analysis.quality_summary.get("ability_category_unknown") == 1
        # ABILITY_ITEMS_REJECTED audit must fire because we dropped a row.
        rejected = _audits(session, AuditEventType.ABILITY_ITEMS_REJECTED)
        assert len(rejected) == 1
        assert rejected[0].summary["abilities_rejected"] == 1

    def test_work_content_missing_for_p_category_blocks_row(self, session):
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        # P ability declared under general_abilities (no work_content parent).
        body = _minimal_record_body()
        body["tasks"][0]["work_contents"][0]["abilities"] = []
        body["tasks"][0]["general_abilities"] = {
            "P": [
                {
                    "ability_code": "P-1.1.1",
                    "ability_major_category_code": "P",
                    "ability_content": "orphaned P",
                }
            ]
        }
        ref = _make_ref_with_payload(session, storage, record_body=body)
        result = ability_writer_write(session, ref, body)
        items = list(session.scalars(select(models.OccupationalAbilityItem)))
        assert items == []
        analysis = session.get(models.OccupationalAbilityAnalysis, result.analysis_id)
        assert analysis.quality_summary.get("work_content_missing_for_p_category") == 1

    def test_task_code_duplicate_keeps_first_only(self, session):
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        body = _minimal_record_body()
        # Append a second task with the same code; only the first should land.
        body["tasks"].append(
            {
                "task_code": "1",
                "task_name": "数据采集 (DUP)",
                "task_description": None,
                "display_order": 2,
                "trace": {},
                "work_contents": [],
                "general_abilities": {},
            }
        )
        ref = _make_ref_with_payload(session, storage, record_body=body)
        result = ability_writer_write(session, ref, body)
        tasks = list(session.scalars(select(models.OccupationalWorkTask)))
        assert len(tasks) == 1
        assert tasks[0].task_name == "数据采集"  # first one wins
        analysis = session.get(models.OccupationalAbilityAnalysis, result.analysis_id)
        assert analysis.quality_summary.get("task_code_duplicate") == 1

    def test_unknown_category_for_general_abilities_rejected(self, session):
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        body = _minimal_record_body()
        body["tasks"][0]["general_abilities"] = {
            "X": [
                {
                    "ability_code": "X-1.1",
                    # category not set inside the entry — writer falls back
                    # to the dict key, then checks against category_schema.
                    "ability_content": "alien category",
                }
            ]
        }
        ref = _make_ref_with_payload(session, storage, record_body=body)
        ability_writer_write(session, ref, body)
        analyses = list(session.scalars(select(models.OccupationalAbilityAnalysis)))
        assert analyses[0].quality_summary.get("ability_category_unknown") == 1


class TestWriterUpsert:
    def test_rerun_replaces_previous_analysis(self, session):
        """Dataset-level upsert (§3.3): re-running the writer for the same
        normalized_ref must delete the old analysis (cascade) and insert
        a fresh one — never accumulate stale rows.
        """
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        body_v1 = _minimal_record_body()
        ref = _make_ref_with_payload(session, storage, record_body=body_v1)

        result_v1 = ability_writer_write(session, ref, body_v1)
        session.flush()
        analyses_after_v1 = list(
            session.scalars(select(models.OccupationalAbilityAnalysis))
        )
        assert len(analyses_after_v1) == 1
        assert analyses_after_v1[0].id == result_v1.analysis_id

        # Re-run with a different body
        body_v2 = _minimal_record_body()
        body_v2["analysis"]["major_name"] = "updated_major"
        body_v2["tasks"][0]["task_name"] = "数据采集 v2"
        result_v2 = ability_writer_write(session, ref, body_v2)
        session.flush()

        analyses_after_v2 = list(
            session.scalars(select(models.OccupationalAbilityAnalysis))
        )
        assert len(analyses_after_v2) == 1, "upsert must collapse to one row"
        assert analyses_after_v2[0].id != result_v1.analysis_id, (
            "upsert must delete-then-insert, not update in place"
        )
        assert analyses_after_v2[0].major_name == "updated_major"

    def test_cascade_clears_children_on_rerun(self, session):
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        body = _minimal_record_body()
        ref = _make_ref_with_payload(session, storage, record_body=body)
        ability_writer_write(session, ref, body)
        # snapshot child ids
        old_task_ids = {t.id for t in session.scalars(select(models.OccupationalWorkTask))}
        # re-run
        ability_writer_write(session, ref, body)
        new_task_ids = {t.id for t in session.scalars(select(models.OccupationalWorkTask))}
        assert new_task_ids.isdisjoint(old_task_ids), "child rows must be replaced, not appended"

    def test_unique_constraint_per_normalized_ref(self, session):
        """The writer's upsert hinges on `uq_oaa_normalized_ref` — verify
        the constraint itself rejects parallel inserts so the writer logic
        stays the only entry point.
        """
        from sqlalchemy.exc import IntegrityError

        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        body = _minimal_record_body()
        ref = _make_ref_with_payload(session, storage, record_body=body)
        ability_writer_write(session, ref, body)
        # bare second insert with the same normalized_ref_id must fail
        profile = session.scalar(select(models.AbilityAnalysisProfile))
        rogue = models.OccupationalAbilityAnalysis(
            normalized_ref_id=ref.id,
            asset_version_id=ref.version_id,
            profile_id=profile.id,
            analysis_model="PGSD",
            schema_version="ability_analysis.pgsd.v1",
        )
        session.add(rogue)
        with pytest.raises(IntegrityError):
            session.flush()
        session.rollback()


class TestWriterSkipPaths:
    def test_missing_profile_returns_skipped(self, session):
        # PGSD is NOT seeded → writer cannot find the profile.
        storage = InMemoryObjectStorage()
        body = _minimal_record_body()
        ref = _make_ref_with_payload(session, storage, record_body=body)
        result = ability_writer_write(session, ref, body)
        assert result.skipped is True
        assert result.reason == "profile_not_found"
        # No analysis row should exist
        rows = list(session.scalars(select(models.OccupationalAbilityAnalysis)))
        assert rows == []

    def test_missing_analysis_block_skips(self, session):
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        ref = _make_ref_with_payload(session, storage, record_body={"tasks": []})
        result = ability_writer_write(session, ref, {"tasks": []})
        assert result.skipped is True
        assert result.reason == "missing_analysis_block"

    def test_analysis_model_mismatch_skips(self, session):
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        body = _minimal_record_body()
        body["analysis"]["analysis_model"] = "NOT_PGSD"
        ref = _make_ref_with_payload(session, storage, record_body=body)
        result = ability_writer_write(session, ref, body)
        assert result.skipped is True
        assert result.reason == "analysis_model_mismatch"


# ===========================================================================
# 3. Dispatcher integration — domain_normalize end-to-end via MinIO payload
# ===========================================================================


class TestDispatcherIntegration:
    def test_dispatcher_routes_to_writer(self, session):
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        body = _minimal_record_body()
        ref = _make_ref_with_payload(session, storage, record_body=body)
        result = dispatch_domain_normalize(session, ref, storage=storage)
        assert result.skipped is False
        assert result.domain_profile == "ability_analysis.pgsd.v1"
        assert result.analysis_id is not None
        # writer side effect: one analysis row.
        rows = list(session.scalars(select(models.OccupationalAbilityAnalysis)))
        assert len(rows) == 1


# ===========================================================================
# 4. End-to-end with sample 2 — build canonical record_body from xlsx
# ===========================================================================


def _build_record_body_from_sample(xlsx_bytes: bytes) -> dict:
    """Minimal sample-aware builder: traverses the per-task sheets of
    sample 2 to construct a canonical record_body. Not part of production
    Pipeline B (B5 / future slice owns the canonical transform); kept
    test-local so the E2E demonstrates the writer can ingest the sample
    when the upstream transform lands.
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    tasks: list[dict] = []
    work_content_count = 0
    ability_count = 0

    for sheet_name in wb.sheetnames:
        # Per-task sheets start with `<digit>.` (e.g. "1.数据采集") — anything
        # else is the overview sheet which we skip for body construction.
        if not sheet_name or not sheet_name[0].isdigit() or "." not in sheet_name:
            continue
        task_code, _, task_name = sheet_name.partition(".")
        ws = wb[sheet_name]

        # Forward-fill the merged-cell first column (大类) which carries
        # P/G/S/D markers across the rows.
        rows: list[tuple] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(tuple("" if v is None else str(v).strip() for v in row))
        if not rows:
            continue

        # Forward-fill across all columns (handles merged cells in any column).
        ffilled: list[list[str]] = []
        last: list[str] = []
        for row in rows:
            row_list = list(row)
            for idx, val in enumerate(row_list):
                if val:
                    if idx >= len(last):
                        last.extend([""] * (idx - len(last) + 1))
                    last[idx] = val
                else:
                    if idx < len(last):
                        row_list[idx] = last[idx]
            ffilled.append(row_list)

        wc_by_code: dict[str, dict] = {}
        general_abilities: dict[str, list[dict]] = {"G": [], "S": [], "D": []}

        # Heuristic: scan every cell for an ability code; the row carrying
        # the code carries the ability_content somewhere else in the same row.
        import re
        pat_p = re.compile(r"^P-\d+\.\d+\.\d+$")
        pat_other = re.compile(r"^[GSD]-\d+\.\d+$")
        for row in ffilled:
            for cell in row:
                if not cell:
                    continue
                match_p = pat_p.match(cell)
                match_other = pat_other.match(cell)
                if not match_p and not match_other:
                    continue
                # ability_content = the longest other cell in this row
                others = [c for c in row if c and c != cell and not pat_p.match(c) and not pat_other.match(c)]
                ability_content = max(others, key=len) if others else cell
                if match_p:
                    # work_content code = first two segments of the ability code
                    body = cell[2:]  # strip "P-"
                    parts = body.split(".")
                    wc_code = ".".join(parts[:2])
                    wc = wc_by_code.setdefault(
                        wc_code,
                        {
                            "content_code": wc_code,
                            "content_name": f"工作内容 {wc_code}",
                            "abilities": [],
                        },
                    )
                    wc["abilities"].append(
                        {
                            "ability_code": cell,
                            "ability_major_category_code": "P",
                            "ability_content": ability_content,
                        }
                    )
                    ability_count += 1
                else:
                    cat = cell[0]
                    general_abilities.setdefault(cat, []).append(
                        {
                            "ability_code": cell,
                            "ability_major_category_code": cat,
                            "ability_content": ability_content,
                        }
                    )
                    ability_count += 1

        work_content_count += len(wc_by_code)
        tasks.append(
            {
                "task_code": task_code,
                "task_name": task_name,
                "task_description": None,
                "display_order": int(task_code) if task_code.isdigit() else 0,
                "trace": {"sheet": sheet_name},
                "work_contents": list(wc_by_code.values()),
                "general_abilities": general_abilities,
            }
        )
    return {
        "analysis": {
            "major_name": "大数据技术应用",
            "major_direction": None,
            "analysis_model": "PGSD",
            "task_count": len(tasks),
            "work_content_count": work_content_count,
            "ability_item_count": ability_count,
        },
        "tasks": tasks,
    }


@pytest.mark.skipif(not SAMPLE_ABILITY_XLSX.exists(), reason="sample 2 missing")
class TestSample2EndToEnd:
    @pytest.fixture
    def state(self, session):
        _seed_pgsd_profile(session)
        storage = InMemoryObjectStorage()
        record_body = _build_record_body_from_sample(SAMPLE_ABILITY_XLSX.read_bytes())
        ref = _make_ref_with_payload(session, storage, record_body=record_body)
        result = dispatch_domain_normalize(session, ref, storage=storage)
        session.flush()
        return result, ref, record_body

    def test_dispatcher_succeeds(self, state):
        result, _, _ = state
        assert result.skipped is False
        assert result.analysis_id is not None

    def test_sample_has_multiple_tasks(self, session, state):
        result, _, _ = state
        tasks = list(
            session.scalars(
                select(models.OccupationalWorkTask).where(
                    models.OccupationalWorkTask.analysis_id == result.analysis_id
                )
            )
        )
        # sample 2 has 4 typical work tasks
        assert len(tasks) >= 3, f"expected >=3 tasks, got {len(tasks)}"

    def test_sample_has_p_abilities(self, session, state):
        result, _, _ = state
        items = list(
            session.scalars(
                select(models.OccupationalAbilityItem).where(
                    models.OccupationalAbilityItem.analysis_id == result.analysis_id,
                    models.OccupationalAbilityItem.ability_major_category_code == "P",
                )
            )
        )
        assert len(items) >= 10, "sample 2 has dozens of P abilities"
        # All P items must have a work_content parent
        for it in items:
            assert it.work_content_id is not None

    def test_sample_has_gsd_abilities_with_null_work_content(self, session, state):
        result, _, _ = state
        gsd_items = list(
            session.scalars(
                select(models.OccupationalAbilityItem).where(
                    models.OccupationalAbilityItem.analysis_id == result.analysis_id,
                    models.OccupationalAbilityItem.ability_major_category_code.in_(
                        ["G", "S", "D"]
                    ),
                )
            )
        )
        assert len(gsd_items) >= 3, "sample 2 has G/S/D rows"
        for it in gsd_items:
            assert it.work_content_id is None

    def test_sample_relations_emitted(self, session, state):
        result, _, _ = state
        relations = list(
            session.scalars(
                select(models.OccupationalAbilityRelation).where(
                    models.OccupationalAbilityRelation.analysis_id == result.analysis_id
                )
            )
        )
        rel_types = {r.relation_type for r in relations}
        assert "TASK_HAS_WORK_CONTENT" in rel_types
        assert "WORK_CONTENT_REQUIRES_ABILITY" in rel_types
        # B6 must NEVER write the other two whitelisted relation types
        assert "ABILITY_DERIVED_FROM_JOB_REQUIREMENT" not in rel_types
        assert "ABILITY_RELATED_TO_SKILL" not in rel_types

    def test_sample_audit_persisted(self, session, state):
        # exactly one of each per dispatch
        persisted = _audits(session, AuditEventType.ABILITY_ANALYSIS_PERSISTED)
        items_persisted = _audits(session, AuditEventType.ABILITY_ITEMS_PERSISTED)
        assert len(persisted) >= 1
        assert len(items_persisted) >= 1
