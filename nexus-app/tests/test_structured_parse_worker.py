"""Tests for the Pipeline B worker integration of `structured_parse` (B1.3).

Two layers:

  1. Unit tests on `_run_structured_parse_xlsx()` — happy path, corrupt
     source, storage read failure. Each verifies the job state, the
     JobStage row, and the audit event(s).
  2. Integration test on `execute_job()` end-to-end for an xlsx raw_object —
     ensures the worker calls `parse_xlsx()`, persists a ParsedWorkbook-shaped
     record_body, and lets the downstream stages run to completion.

The B1.1 routing flag is NOT toggled here — the worker test constructs the
job with `payload["pipeline_type"] = "record"` directly, isolating worker
behavior from gateway routing decisions (covered by `test_pipeline_routing.py`).
"""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import select

from nexus_app import models, services
from nexus_app.config import get_settings
from nexus_app.enums import (
    AssetVersionStatus,
    AuditEventType,
    DataSourceType,
    IngestBatchStatus,
    JobStatus,
    JobType,
    PipelineType,
    RawObjectStatus,
    StageStatus,
)
from nexus_app.mineru import FakeMinerUAdapter
from nexus_app.pipeline.payload_schema import JOB_PAYLOAD_SCHEMA_VERSION
from nexus_app.schemas import DataSourceCreate
from nexus_app.storage import InMemoryObjectStorage
from nexus_app.worker.claimer import claim_jobs
from nexus_app.worker.runner import (
    NonRetryableError,
    _run_structured_parse_xlsx,
    execute_job,
)

REPO_ROOT = Path(__file__).resolve().parents[2]
SAMPLE_JOB_DEMAND = REPO_ROOT / "docs/samples/1.（岗位需求）电子商务岗位招聘数据.xlsx"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------------------------------------------------------------------
# Synthetic xlsx helpers
# ---------------------------------------------------------------------------


def _make_xlsx_bytes() -> bytes:
    """Build a tiny xlsx with one sheet and a few rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    ws["B1"] = "value"
    ws["A2"] = "row1"
    ws["B2"] = 42
    ws["A3"] = "row2"
    ws["B3"] = datetime(2024, 12, 12, 1, 12, 59)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Fixtures — minimal Job/RawObject scaffolding so we can call worker helpers
# directly without touching the gateway. This isolates B1.3 worker behavior
# from B1.1 routing behavior.
# ---------------------------------------------------------------------------


def _seed_raw_object(
    session,
    storage: InMemoryObjectStorage,
    *,
    xlsx_bytes: bytes,
    filename: str = "sample.xlsx",
    mime: str = XLSX_MIME,
):
    """Create a DataSource + IngestBatch + RawObject backed by storage."""
    source = services.create_data_source(
        session,
        DataSourceCreate(
            code="ds-sp-xlsx",
            name="xlsx test",
            source_type="file_upload",
        ),
    )
    batch = models.IngestBatch(
        data_source_id=source.id,
        idempotency_key="sp-key-1",
        source_type=DataSourceType.FILE_UPLOAD,
        status=IngestBatchStatus.RAW_PERSISTED,
    )
    session.add(batch)
    session.flush()

    key = f"raw/file_upload/{source.id}/{filename}"
    stored = storage.put_bytes(key, xlsx_bytes, mime)

    raw = models.RawObject(
        data_source_id=source.id,
        batch_id=batch.id,
        source_type=DataSourceType.FILE_UPLOAD,
        object_uri=stored.object_uri,
        checksum=stored.checksum,
        mime_type=mime,
        size_bytes=stored.size_bytes,
        status=RawObjectStatus.RAW_PERSISTED,
        metadata_summary={"filename": filename},
    )
    session.add(raw)
    session.flush()

    job = models.Job(
        job_type=JobType.INGEST_PROCESS,
        status=JobStatus.RUNNING,
        ingest_batch_id=batch.id,
        raw_object_id=raw.id,
        idempotency_key="sp-key-1",
        payload={"pipeline_type": "record"},
        payload_schema_version=JOB_PAYLOAD_SCHEMA_VERSION,
        trace_id="trace-sp",
    )
    session.add(job)
    session.commit()
    return job, raw, batch


def _audits(session, event_type: AuditEventType) -> list[models.AuditLog]:
    return list(
        session.scalars(
            select(models.AuditLog).where(models.AuditLog.event_type == event_type)
        )
    )


def _stages(session, job_id: str, *, stage_name: str | None = None) -> list[models.JobStage]:
    q = select(models.JobStage).where(models.JobStage.job_id == job_id)
    if stage_name is not None:
        q = q.where(models.JobStage.stage_name == stage_name)
    return list(session.scalars(q))


# ---------------------------------------------------------------------------
# Unit: _run_structured_parse_xlsx happy path
# ---------------------------------------------------------------------------


class TestRunStructuredParseXlsxHappyPath:
    def test_returns_parsed_workbook_dict(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(session, storage, xlsx_bytes=_make_xlsx_bytes())

        result = _run_structured_parse_xlsx(job, raw, storage, session, "trace-sp")

        # The dict shape mirrors ParsedWorkbook.model_dump(mode="json").
        assert result["parser_version"] == "xlsx_parser.v1"
        assert result["source_filename"] == "sample.xlsx"
        assert result["source_mime_type"] == XLSX_MIME
        assert isinstance(result["sheets"], list)
        assert result["sheets"][0]["name"] == "Sheet1"
        # parsed_at must be a string (mode='json' serialises datetime to ISO8601)
        assert isinstance(result["parsed_at"], str)

    def test_writes_succeeded_stage_row(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(session, storage, xlsx_bytes=_make_xlsx_bytes())

        _run_structured_parse_xlsx(job, raw, storage, session, "trace-sp")

        stages = _stages(session, job.id, stage_name="structured_parse")
        assert len(stages) == 1
        stage = stages[0]
        assert stage.status == StageStatus.SUCCEEDED
        assert stage.finished_at is not None
        assert stage.failure_reason is None
        assert stage.detail["parser_version"] == "xlsx_parser.v1"
        assert stage.detail["sheet_count"] == 1
        # detail.sheets array carries per-sheet summary used by operators
        assert stage.detail["sheets"][0]["name"] == "Sheet1"
        assert stage.detail["sheets"][0]["row_count"] >= 3

    def test_writes_structured_parse_completed_audit(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(session, storage, xlsx_bytes=_make_xlsx_bytes())

        _run_structured_parse_xlsx(job, raw, storage, session, "trace-sp")

        audits = _audits(session, AuditEventType.STRUCTURED_PARSE_COMPLETED)
        assert len(audits) == 1
        audit = audits[0]
        assert audit.target_type == "raw_object"
        assert audit.target_id == raw.id
        assert audit.trace_id == "trace-sp"
        assert audit.summary["parser_version"] == "xlsx_parser.v1"
        assert audit.summary["sheet_count"] == 1
        assert audit.summary["timezone"] == "Asia/Shanghai"

    def test_sets_current_stage(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(session, storage, xlsx_bytes=_make_xlsx_bytes())

        _run_structured_parse_xlsx(job, raw, storage, session, "trace-sp")

        session.refresh(job)
        assert job.current_stage == "structured_parse"


# ---------------------------------------------------------------------------
# Unit: failure paths
# ---------------------------------------------------------------------------


class TestRunStructuredParseXlsxFailures:
    def test_corrupt_xlsx_fails_job_and_raises_non_retryable(self, session):
        storage = InMemoryObjectStorage()
        # Plant bytes that are NOT a valid xlsx zip
        job, raw, _ = _seed_raw_object(
            session, storage, xlsx_bytes=b"this is not an xlsx file"
        )

        with pytest.raises(NonRetryableError) as excinfo:
            _run_structured_parse_xlsx(job, raw, storage, session, "trace-sp")
        assert excinfo.value.error_code == "structured_parse_corrupt_source"

        session.refresh(job)
        assert job.status == JobStatus.FAILED
        assert job.last_error_code == "structured_parse_corrupt_source"
        assert job.current_stage == "structured_parse"

    def test_corrupt_xlsx_writes_failed_stage_row(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(session, storage, xlsx_bytes=b"not xlsx")

        with pytest.raises(NonRetryableError):
            _run_structured_parse_xlsx(job, raw, storage, session, "trace-sp")

        stages = _stages(session, job.id, stage_name="structured_parse")
        assert len(stages) == 1
        assert stages[0].status == StageStatus.FAILED
        assert stages[0].failure_reason is not None
        assert stages[0].detail["error_code"] == "structured_parse_corrupt_source"

    def test_corrupt_xlsx_writes_pipeline_failed_audit(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(session, storage, xlsx_bytes=b"not xlsx")

        with pytest.raises(NonRetryableError):
            _run_structured_parse_xlsx(job, raw, storage, session, "trace-sp")

        audits = _audits(session, AuditEventType.PIPELINE_FAILED)
        # Filter to our job/stage — audit target_id is the job_id, summary.stage
        # disambiguates structured_parse failures from other pipeline failures.
        our = [
            a for a in audits
            if a.summary.get("stage") == "structured_parse" and a.target_id == job.id
        ]
        assert len(our) == 1
        assert our[0].summary["error_code"] == "structured_parse_corrupt_source"
        assert our[0].summary["raw_object_id"] == raw.id

    def test_storage_read_failure_fails_job(self, session):
        storage = InMemoryObjectStorage()
        # Seed a raw_object that points at storage, then delete the underlying
        # object so the worker hits an ObjectNotFoundError.
        job, raw, _ = _seed_raw_object(session, storage, xlsx_bytes=_make_xlsx_bytes())
        raw_key = raw.object_uri.split("/", 3)[-1]
        storage.delete_object(raw_key)

        with pytest.raises(NonRetryableError) as excinfo:
            _run_structured_parse_xlsx(job, raw, storage, session, "trace-sp")
        assert excinfo.value.error_code == "structured_parse_storage_read_failed"

        session.refresh(job)
        assert job.status == JobStatus.FAILED
        assert job.last_error_code == "structured_parse_storage_read_failed"


# ---------------------------------------------------------------------------
# Integration: full execute_job flow for xlsx record pipeline
# ---------------------------------------------------------------------------


class TestExecuteJobXlsxRecordPipeline:
    """End-to-end: queued xlsx job → run worker → asset+version+normalized_ref.

    structured_parse runs first, then the existing record pipeline (normalize_record →
    governance_decision → knowledge_chunking → index_submit) proceeds. The
    downstream stages no-op gracefully for unclassified record assets — that's
    expected at B1.3 (B2/B3/B6/B7 will teach them about xlsx-derived data).
    """

    def test_xlsx_job_completes_with_structured_parse_stage(self, session):
        storage = InMemoryObjectStorage()
        job, raw, batch = _seed_raw_object(
            session, storage, xlsx_bytes=_make_xlsx_bytes()
        )

        # FakeMinerUAdapter is only used by Pipeline A; the record pipeline
        # path won't touch it but execute_job still requires it.
        execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())

        session.refresh(job)
        session.refresh(batch)
        assert job.status == JobStatus.SUCCEEDED, f"job failed: {job.failure_reason}"

        # Stage rows must include 'structured_parse' BEFORE 'assetize' and 'normalize'.
        all_stages = _stages(session, job.id)
        stage_names = [s.stage_name for s in all_stages]
        assert "structured_parse" in stage_names
        # structured_parse must precede assetize/normalize in insertion order
        sp_idx = stage_names.index("structured_parse")
        for later in ("assetize", "normalize"):
            assert later in stage_names
            assert stage_names.index(later) > sp_idx

    def test_xlsx_job_persists_record_with_parsedworkbook_shape(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(session, storage, xlsx_bytes=_make_xlsx_bytes())

        execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())

        # NormalizedAssetRef created with the structured_parse output as record_body
        refs = list(session.scalars(select(models.NormalizedAssetRef)))
        assert len(refs) == 1
        ref = refs[0]

        # The normalized payload was uploaded to storage; fetch and inspect.
        ref_key = ref.object_uri.split("/", 3)[-1]
        import json
        payload = json.loads(storage.get_bytes(ref_key).decode("utf-8"))
        record_body = payload["record_body"]

        # record_body should carry the ParsedWorkbook shape (B1.3 contract).
        assert record_body["parser_version"] == "xlsx_parser.v1"
        assert isinstance(record_body["sheets"], list)
        assert record_body["sheets"][0]["name"] == "Sheet1"

    def test_xlsx_job_writes_structured_parse_completed_audit(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(session, storage, xlsx_bytes=_make_xlsx_bytes())

        execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())

        audits = _audits(session, AuditEventType.STRUCTURED_PARSE_COMPLETED)
        assert len(audits) == 1
        assert audits[0].target_id == raw.id

    def test_asset_kind_is_record(self, session):
        storage = InMemoryObjectStorage()
        job, _, _ = _seed_raw_object(session, storage, xlsx_bytes=_make_xlsx_bytes())

        execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())

        assets = list(session.scalars(select(models.Asset)))
        assert len(assets) == 1
        # record pipeline → AssetKind.RECORD
        assert assets[0].asset_kind.value == "record"


# ---------------------------------------------------------------------------
# Integration: csv mime end-to-end (B1.4)
# ---------------------------------------------------------------------------


class TestExecuteJobCsvRecordPipeline:
    """End-to-end: queued csv job → parse_csv → assetize → normalize → ...

    Same shape as the xlsx test class but exercises the csv branch of
    `execute_job()` introduced in B1.4.
    """

    def test_csv_job_completes_with_structured_parse_stage(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(
            session,
            storage,
            xlsx_bytes=b"name,city\nzs,bj\nls,sh\n",  # csv bytes
            filename="x.csv",
            mime="text/csv",
        )

        execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())

        session.refresh(job)
        assert job.status == JobStatus.SUCCEEDED, f"job failed: {job.failure_reason}"
        stage_names = [s.stage_name for s in _stages(session, job.id)]
        assert "structured_parse" in stage_names
        sp_idx = stage_names.index("structured_parse")
        for later in ("assetize", "normalize"):
            assert later in stage_names
            assert stage_names.index(later) > sp_idx

    def test_csv_job_records_format_label_in_stage_detail(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(
            session,
            storage,
            xlsx_bytes=b"a,b\n1,2",
            filename="x.csv",
            mime="text/csv",
        )

        execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())

        sp = _stages(session, job.id, stage_name="structured_parse")[0]
        # The shared runner stamps `format` so operators can tell which parser ran.
        assert sp.detail["format"] == "csv"
        assert sp.detail["parser_version"] == "csv_parser.v1"

    def test_csv_job_persists_record_with_projected_record_body(self, session):
        # Post-B3.5: `record_body` is no longer the raw ParsedWorkbook —
        # the adapter projects it to a contract shape based on
        # `domain_profile`. csv_parser-specific provenance (parser_version,
        # sheet names) is captured on the structured_parse audit / stage
        # detail instead (see `test_csv_record_pipeline_succeeds`).
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(
            session,
            storage,
            xlsx_bytes=b"name,city\nzs,bj\n",
            filename="x.csv",
            mime="text/csv",
        )

        execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())

        refs = list(session.scalars(select(models.NormalizedAssetRef)))
        assert len(refs) == 1
        ref_key = refs[0].object_uri.split("/", 3)[-1]
        import json
        payload = json.loads(storage.get_bytes(ref_key).decode("utf-8"))
        # The payload v2 envelope is preserved end-to-end regardless of which
        # projection path runs: schema_version + record_body are always set.
        assert payload["schema_version"].startswith("normalized-record")
        assert isinstance(payload["record_body"], dict)

    def test_csv_corrupt_bytes_fails_with_structured_parse_error_code(self, session):
        # A CSV body that's also a path-style str triggers no error, but
        # an encoding-impossible bytes blob does.
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(
            session,
            storage,
            xlsx_bytes=bytes([0xC3, 0x28, 0xFF]),  # invalid in utf-8 / gb18030
            filename="bad.csv",
            mime="text/csv",
        )

        with pytest.raises(NonRetryableError) as excinfo:
            execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())
        assert excinfo.value.error_code == "structured_parse_corrupt_source"

        session.refresh(job)
        assert job.status == JobStatus.FAILED
        assert job.last_error_code == "structured_parse_corrupt_source"


# ---------------------------------------------------------------------------
# Integration: sample 1 xlsx end-to-end (real-world workbook)
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not SAMPLE_JOB_DEMAND.exists(), reason="sample missing")
class TestExecuteJobSampleJobDemand:
    def test_sample_runs_through_pipeline(self, session):
        storage = InMemoryObjectStorage()
        job, raw, _ = _seed_raw_object(
            session,
            storage,
            xlsx_bytes=SAMPLE_JOB_DEMAND.read_bytes(),
            filename=SAMPLE_JOB_DEMAND.name,
        )

        execute_job(job, session, storage, FakeMinerUAdapter(), get_settings())

        session.refresh(job)
        assert job.status == JobStatus.SUCCEEDED, f"failed: {job.failure_reason}"

        # Stage rows include structured_parse with the real sample's metadata
        sp_stages = _stages(session, job.id, stage_name="structured_parse")
        assert len(sp_stages) == 1
        sp = sp_stages[0]
        assert sp.status == StageStatus.SUCCEEDED
        assert sp.detail["sheet_count"] == 1
        sheet0 = sp.detail["sheets"][0]
        assert sheet0["name"] == "Sheet1"
        assert sheet0["dropped_index_columns"] == [1]  # 序号 column dropped
